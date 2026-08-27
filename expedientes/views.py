"""Vistas de GDE — Gestión Digital de Expedientes."""

import mimetypes
import os
from decimal import Decimal
from urllib.parse import quote

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.http import content_disposition_header, url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST

from . import documentos as generador
from . import pdf as conversor
from .auditoria import registrar
from .compresion import NoSePudoComprimir, comprimir_imagen, comprimir_pdf
from .completitud import completitud
from .forms import (
    BonoExtraForm, DatosContratacionForm, DocumentoForm, EscaneoForm,
    FiltroTrabajadorForm,
    HijoForm, RemuneracionForm, TrabajadorForm,
)
from .models import (
    AsignacionPago, ConceptoPago, DatosContratacion, Documento, Hijo,
    RegistroAuditoria, Trabajador,
)
from .permisos import (
    exigir_borrar,
    exigir_borrar_del_expediente,
    exigir_editar_trabajador,
    exigir_gestionar_pagos,
    exigir_ver_trabajador,
    puede_gestionar_pagos,
    trabajadores_visibles,
)


@login_required
def panel(request):
    """Tablero con métricas según el alcance del usuario."""
    trabajadores = trabajadores_visibles(request.user)
    docs = Documento.objects.filter(activo=True, trabajador__in=trabajadores)
    hoy = timezone.localdate()
    limite = hoy + timezone.timedelta(days=30)

    por_vencer = docs.filter(
        fecha_vencimiento__isnull=False,
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=limite,
    ).select_related("trabajador", "tipo").order_by("fecha_vencimiento")[:10]

    contexto = {
        "total_trabajadores": trabajadores.count(),
        "total_documentos": docs.count(),
        "por_vencer": por_vencer,
        "cantidad_por_vencer": docs.filter(
            fecha_vencimiento__isnull=False,
            fecha_vencimiento__gte=hoy,
            fecha_vencimiento__lte=limite,
        ).count(),
        "vencidos": docs.filter(fecha_vencimiento__lt=hoy).count(),
    }
    return render(request, "expedientes/panel.html", contexto)


def _querystring_sin_pagina(request):
    """Los filtros actuales, sin `page`, para reusarlos en la paginación.

    Sin esto, pasar a la página 2 perdía todos los filtros: se volvía al
    listado completo. Con varias tiendas marcadas se notaba todavía más.
    """
    parametros = request.GET.copy()
    parametros.pop("page", None)
    return parametros.urlencode()


@login_required
def trabajador_list(request):
    form = FiltroTrabajadorForm(request.GET or None, usuario=request.user)
    qs = trabajadores_visibles(request.user).select_related(
        "sede", "sede__zona", "puesto", "tipo_documento"
    ).annotate(
        cant_docs=Count("documentos", filter=Q(documentos__activo=True))
    ).order_by("apellidos", "nombres")

    if form.is_valid():
        q = form.cleaned_data.get("q")
        sedes = form.cleaned_data.get("sedes")
        estado = form.cleaned_data.get("estado")
        docs = form.cleaned_data.get("docs")
        if q:
            qs = qs.filter(
                Q(nombres__icontains=q)
                | Q(apellidos__icontains=q)
                | Q(documento_identidad__icontains=q)
            )
        if sedes:
            qs = qs.filter(sede__in=sedes)
        if estado:
            qs = qs.filter(estado=estado)
        if docs not in (None, ""):
            qs = qs.filter(cant_docs=int(docs))
        # `creado_en` es fecha-hora: se compara por día para que «hasta»
        # incluya todo el día elegido.
        creado_desde = form.cleaned_data.get("creado_desde")
        creado_hasta = form.cleaned_data.get("creado_hasta")
        ingreso_desde = form.cleaned_data.get("ingreso_desde")
        ingreso_hasta = form.cleaned_data.get("ingreso_hasta")
        if creado_desde:
            qs = qs.filter(creado_en__date__gte=creado_desde)
        if creado_hasta:
            qs = qs.filter(creado_en__date__lte=creado_hasta)
        if ingreso_desde:
            qs = qs.filter(fecha_ingreso__gte=ingreso_desde)
        if ingreso_hasta:
            qs = qs.filter(fecha_ingreso__lte=ingreso_hasta)

    paginador = Paginator(qs, 15)
    pagina = paginador.get_page(request.GET.get("page"))

    contexto = {"form": form, "pagina": pagina,
                "querystring": _querystring_sin_pagina(request),
                # El «Limpiar filtros» solo se ofrece cuando hay algo que
                # limpiar (sin contar la paginación).
                "filtros_activos": any(k != "page" for k in request.GET)}
    # Respuesta parcial para HTMX (búsqueda en vivo).
    if request.headers.get("HX-Request"):
        return render(request, "expedientes/_tabla_trabajadores.html", contexto)
    return render(request, "expedientes/trabajador_list.html", contexto)


@login_required
def trabajador_detail(request, pk):
    trabajador = get_object_or_404(
        Trabajador.objects.select_related("sede", "sede__zona",
                                         "tipo_documento"), pk=pk
    )
    exigir_ver_trabajador(request.user, trabajador)
    registrar(request, RegistroAuditoria.Accion.VER, entidad="Trabajador",
              objeto_id=trabajador.pk, descripcion=f"Consultó expediente de {trabajador}")

    documentos = (
        trabajador.documentos.filter(activo=True)
        .select_related("tipo", "subido_por")
        .order_by("tipo__orden", "-subido_en")
    )
    cargados, requeridos, faltantes = trabajador.estado_completitud()

    contexto = {
        "trabajador": trabajador,
        "hijos": list(trabajador.hijos.all()),
        "form_hijo": HijoForm(trabajador=trabajador),
        "documentos": documentos,
        "form_doc": DocumentoForm(),
        "cargados": cargados,
        "requeridos": requeridos,
        "faltantes": faltantes,
        "pct_completitud": int(cargados / requeridos * 100) if requeridos else 100,
        # El navegador avisa que un archivo pesa demasiado antes de empezar a
        # subirlo. Para eso necesita el mismo número que usa el servidor.
        "documento_max_bytes": settings.DOCUMENTOS_MAX_BYTES,
        "documento_max_mb": settings.DOCUMENTOS_MAX_BYTES // 1024 // 1024,
    }

    # Remuneración: solo Admin y RRHH Interior de la zona.
    contexto["puede_pagos"] = puede_gestionar_pagos(request.user, trabajador)
    if contexto["puede_pagos"]:
        pagos = list(
            trabajador.pagos.filter(activo=True).select_related("concepto", "moneda")
        )
        contexto["pagos"] = pagos
        contexto["totales_pagos"] = _totales_por_moneda(pagos)
        # Grilla con los conceptos del catálogo + los bonos extras cargados a mano.
        contexto["form_remuneracion"] = RemuneracionForm(trabajador=trabajador)
        contexto["form_extra"] = BonoExtraForm()
        contexto["extras"] = [p for p in pagos if p.concepto_id is None]

        # Documentos corporativos: se arman con los mismos datos del expediente.
        contexto["plantillas"] = [
            {"clave": clave, "titulo": meta["titulo"]}
            for clave, meta in generador.PLANTILLAS.items()
        ]
        contexto["campos_incompletos"] = generador.campos_incompletos(trabajador)
        # Sin Word en el servidor no hay PDF: se ofrece solo la descarga Word.
        contexto["hay_pdf"] = conversor.hay_conversor()

    return render(request, "expedientes/trabajador_detail.html", contexto)


@login_required
def documento_generar(request, pk, clave):
    """Entrega un documento corporativo en Word, en PDF o listo para imprimir.

    El formato viaja en `?formato=`:
      - `word` (o nada): el .docx/.rtf, para editarlo antes de firmar.
      - `pdf`: el mismo documento convertido, para archivar o mandar por mail.
      - `imprimir`: el PDF **en el navegador**, así se imprime sin bajar nada.

    El PDF lo arma Word en la máquina del servidor, así que sale idéntico al
    documento. Si esa máquina no tiene Word, se avisa y se ofrece el Word.
    """
    formato = request.GET.get("formato", "word")
    if clave not in generador.PLANTILLAS or formato not in {"word", "pdf", "imprimir"}:
        raise Http404("Documento desconocido.")

    trabajador = get_object_or_404(
        Trabajador.objects.select_related("sede", "sede__zona", "puesto",
                                         "tipo_documento"), pk=pk
    )
    exigir_gestionar_pagos(request.user, trabajador)

    try:
        contenido, nombre, _faltantes = generador.generar(clave, trabajador)
    except FileNotFoundError as e:
        messages.error(request, str(e))
        return redirect("expedientes:trabajador_detail", pk=trabajador.pk)

    titulo = generador.PLANTILLAS[clave]["titulo"]

    if formato == "word":
        tipo = ("application/rtf" if nombre.lower().endswith(".rtf") else
                "application/vnd.openxmlformats-officedocument."
                "wordprocessingml.document")
        adjunto = True
    else:
        try:
            contenido = conversor.convertir_a_pdf(contenido, nombre)
        except conversor.ConversionNoDisponible as e:
            messages.error(request, f"{e}")
            return redirect("expedientes:trabajador_detail", pk=trabajador.pk)
        nombre = conversor.nombre_pdf(nombre)
        tipo = "application/pdf"
        # Para imprimir se muestra dentro del navegador: el visor de PDF ya
        # trae su botón de imprimir y no deja archivos sueltos en Descargas.
        adjunto = formato == "pdf"

    accion = {"word": "Generó", "pdf": "Generó en PDF",
              "imprimir": "Abrió para imprimir"}[formato]
    registrar(request, RegistroAuditoria.Accion.DESCARGAR, entidad="Documento generado",
              objeto_id=trabajador.pk,
              descripcion=f"{accion} '{titulo}' de {trabajador}")

    respuesta = HttpResponse(contenido, content_type=tipo)
    disposicion = "attachment" if adjunto else "inline"
    respuesta["Content-Disposition"] = (
        f'{disposicion}; filename="{nombre}"; '
        f"filename*=UTF-8''{quote(nombre)}"
    )
    return respuesta


def _totales_por_moneda(pagos):
    """Totaliza los montos agrupándolos por moneda.

    Las monedas NO se suman entre sí: el sistema no guarda tasas de cambio, así
    que cada moneda se totaliza por separado (180,00 Bs · 400,00 $ · 50,00 €).
    """
    acumulado = {}
    for p in pagos:
        entrada = acumulado.setdefault(
            p.moneda_id, {"moneda": p.moneda, "total": Decimal("0")}
        )
        entrada["total"] += p.monto
    ordenados = sorted(acumulado.values(), key=lambda e: (e["moneda"].orden, e["moneda"].codigo))
    return [
        {"moneda": e["moneda"], "total": e["moneda"].formatear(e["total"])}
        for e in ordenados
    ]


def _errores_legibles(form):
    return "; ".join(
        f"{form.fields[c].label if c in form.fields else c}: {', '.join(e)}"
        for c, e in form.errors.items()
    )


@login_required
@require_POST
def remuneracion_guardar(request, pk):
    """Guarda la grilla completa: un monto por cada concepto del catálogo."""
    trabajador = get_object_or_404(Trabajador.objects.select_related("sede"), pk=pk)
    exigir_gestionar_pagos(request.user, trabajador)

    form = RemuneracionForm(request.POST, trabajador=trabajador)
    if form.is_valid():
        altas, cambios, bajas, bloqueadas = form.guardar(request.user)
        if altas or cambios or bajas:
            registrar(request, RegistroAuditoria.Accion.EDITAR, entidad="Remuneración",
                      objeto_id=trabajador.pk,
                      descripcion=f"Actualizó la remuneración de {trabajador} "
                                  f"({altas} alta/s, {cambios} cambio/s, {bajas} baja/s)")
            messages.success(request, "Remuneración guardada.")
        elif not bloqueadas:
            messages.info(request, "No hubo cambios en la remuneración.")
        if bloqueadas:
            messages.warning(
                request,
                "Quitar un concepto del expediente lo hace el Administrador. "
                f"Se dejó como estaba: {', '.join(bloqueadas)}."
            )
    else:
        messages.error(request, f"Revisá los montos. {_errores_legibles(form)}")
    return redirect("expedientes:trabajador_detail", pk=trabajador.pk)


@login_required
@require_POST
def pago_agregar(request, pk):
    """Agrega un bono extra (fuera del catálogo) al expediente."""
    trabajador = get_object_or_404(
        Trabajador.objects.select_related("sede"), pk=pk
    )
    exigir_gestionar_pagos(request.user, trabajador)

    form = BonoExtraForm(request.POST)
    if form.is_valid():
        pago = form.save(commit=False)
        pago.trabajador = trabajador
        pago.concepto = None  # es un bono libre, no sale del catálogo
        pago.creado_por = request.user
        pago.save()
        registrar(request, RegistroAuditoria.Accion.CREAR, entidad="Remuneración",
                  objeto_id=pago.pk,
                  descripcion=f"Agregó el bono extra '{pago.etiqueta}' por "
                              f"{pago.monto_formateado} a {trabajador}")
        messages.success(request, f"Agregado: {pago.etiqueta} — {pago.monto_formateado}.")
    else:
        messages.error(request, f"No se pudo agregar el bono. {_errores_legibles(form)}")
    return redirect("expedientes:trabajador_detail", pk=trabajador.pk)


@login_required
def pago_editar(request, pk):
    """Edita un bono extra. Los del catálogo se editan en la grilla."""
    pago = get_object_or_404(
        AsignacionPago.objects.select_related("trabajador", "trabajador__sede",
                                              "concepto", "moneda"),
        pk=pk, activo=True,
    )
    exigir_gestionar_pagos(request.user, pago.trabajador)

    if pago.concepto_id is not None:
        messages.info(request, "Los conceptos del catálogo se editan en la grilla "
                               "de remuneración.")
        return redirect("expedientes:trabajador_detail", pk=pago.trabajador_id)

    form = BonoExtraForm(request.POST or None, instance=pago)
    if request.method == "POST" and form.is_valid():
        anterior = f"{pago.etiqueta} — {pago.monto_formateado}"
        pago = form.save()
        registrar(request, RegistroAuditoria.Accion.EDITAR, entidad="Remuneración",
                  objeto_id=pago.pk,
                  descripcion=f"Modificó '{anterior}' → '{pago.etiqueta} — "
                              f"{pago.monto_formateado}' de {pago.trabajador}")
        messages.success(request, "Bono actualizado.")
        return redirect("expedientes:trabajador_detail", pk=pago.trabajador_id)

    return render(request, "expedientes/pago_form.html",
                  {"form": form, "pago": pago, "trabajador": pago.trabajador})


@login_required
@require_POST
def pago_borrar(request, pk):
    """Da de baja un monto (borrado lógico, como los documentos).

    Quitar es del Administrador. RRHH Interior puede cargar y corregir montos,
    pero sacar un concepto del expediente ya no.
    """
    pago = get_object_or_404(
        AsignacionPago.objects.select_related("trabajador", "trabajador__sede",
                                              "concepto", "moneda"),
        pk=pk,
    )
    exigir_gestionar_pagos(request.user, pago.trabajador)
    exigir_borrar_del_expediente(request.user, pago.trabajador)

    pago.activo = False
    pago.save(update_fields=["activo", "actualizado_en"])
    registrar(request, RegistroAuditoria.Accion.BORRAR, entidad="Remuneración",
              objeto_id=pago.pk,
              descripcion=f"Dio de baja '{pago.etiqueta}' ({pago.monto_formateado}) "
                          f"de {pago.trabajador}")
    messages.warning(request, f"Dado de baja: {pago.etiqueta}.")
    return redirect("expedientes:trabajador_detail", pk=pago.trabajador_id)


def _secciones_expediente(form, form_contrato):
    """Agrupa los campos de ambos formularios en las secciones de la pantalla.

    Lo que sobre cae en "Otros datos": así, si mañana se agrega un campo y se
    olvida ubicarlo acá, sigue apareciendo en vez de desaparecer sin aviso.
    """
    reparto = [
        ("Datos personales",
         # La tupla dice "estos dos van juntos en una sola casilla": el tipo
         # es una letra y no merece media pantalla, y separado de la cédula se
         # lee como otro dato en vez de como su prefijo.
         [("tipo_documento", "documento_identidad"), "rif", "nombres", "apellidos",
          "fecha_nacimiento", "telefono", "email"],
         ["ciudad_nacimiento", "estado_civil", "direccion"],
         "La edad y el día/mes/año de nacimiento se calculan solos a partir "
         "de la fecha."),
        ("Puesto y contrato",
         ["sede", "departamento", "puesto", "fecha_ingreso", "estado"],
         ["duracion_dias", "fecha_culminacion", "motivo_contratacion",
          "horario", "ciudad_firma"],
         "Cargá la duración o la fecha de fin: el sistema completa la otra. "
         "La dirección de la tienda sale de la tienda elegida."),
        ("Datos bancarios",
         [],
         ["banco", "prefijo", "numero_cuenta"],
         "La cuenta completa se arma sola con el prefijo y el número."),
        ("Tallas de uniforme",
         [],
         ["talla_camisa", "talla_pantalon", "talla_zapato"],
         "Se usan para el pedido de dotación. No salen en los documentos Word."),
        ("Seguimiento",
         [],
         ["observaciones", "responsable"],
         "Si dejás el responsable vacío se guarda tu usuario."),
    ]

    usados = {"ficha": set(), "contrato": set()}
    secciones = []
    for titulo, de_ficha, de_contrato, nota in reparto:
        campos = []
        for nombre in de_ficha:
            if isinstance(nombre, tuple):
                juntos = [form[n] for n in nombre if n in form.fields]
                usados["ficha"].update(n for n in nombre if n in form.fields)
                if len(juntos) > 1:
                    campos.append({"pareja": juntos})
                elif juntos:
                    campos.append(juntos[0])
            elif nombre in form.fields:
                campos.append(form[nombre])
                usados["ficha"].add(nombre)
        for nombre in de_contrato:
            if nombre in form_contrato.fields:
                campos.append(form_contrato[nombre])
                usados["contrato"].add(nombre)
        if campos:
            secciones.append({"titulo": titulo, "campos": campos, "nota": nota})

    sobrantes = ([form[n] for n in form.fields if n not in usados["ficha"]]
                 + [form_contrato[n] for n in form_contrato.fields
                    if n not in usados["contrato"]])
    if sobrantes:
        secciones.append({"titulo": "Otros datos", "campos": sobrantes, "nota": ""})
    return secciones


def _problemas_del_expediente(secciones, form, form_contrato):
    """Todo lo que impidió guardar, junto y arriba de todo.

    El formulario del expediente son cinco tarjetas largas. Cuando un dato no
    valida, el navegador vuelve al principio de la página y el aviso queda
    enterrado tres pantallas más abajo: se ve el mismo formulario de siempre,
    igualito, y parece que el botón Guardar no hizo nada.

    Se arma en el orden en que están en la pantalla, y cada uno enlaza a su
    campo, así se llega de un toque.
    """
    problemas = []
    for formulario in (form, form_contrato):
        for mensaje in formulario.non_field_errors():
            problemas.append({"etiqueta": "", "ancla": "", "seccion": "",
                              "mensaje": mensaje})
    for seccion in secciones:
        for campo in _uno_por_uno(seccion["campos"]):
            if campo.errors:
                problemas.append({
                    "etiqueta": campo.label,
                    "ancla": campo.id_for_label,
                    "seccion": seccion["titulo"],
                    "mensaje": " ".join(campo.errors),
                })
    return problemas


def _uno_por_uno(campos):
    """Abre las parejas: dos campos en una casilla siguen siendo dos campos.

    Sin esto, un error en la cédula quedaba fuera del resumen de arriba —que es
    justo el que existe para que ningún error quede escondido—.
    """
    for campo in campos:
        if isinstance(campo, dict):
            yield from campo["pareja"]
        else:
            yield campo


def _guardar_expediente(request, form, form_contrato, trabajador=None):
    """Valida y guarda ficha + datos de contratación como una sola carga.

    Devuelve el trabajador guardado, o None si algo no validó. Las dos
    validaciones se corren siempre (sin cortocircuito) para que la pantalla
    muestre juntos todos los errores.
    """
    ok_ficha = form.is_valid()
    ok_contrato = form_contrato.is_valid()
    if not (ok_ficha and ok_contrato):
        return None

    # La fecha de ingreso está en la ficha, pero la necesitan los dos.
    fecha_ingreso = form.cleaned_data.get("fecha_ingreso")
    if not form_contrato.validar_contra_ingreso(fecha_ingreso):
        return None

    trabajador = form.save()
    datos = form_contrato.save(commit=False)
    datos.trabajador = trabajador
    datos.actualizado_por = request.user
    if not datos.responsable:
        datos.responsable = request.user.get_full_name() or request.user.get_username()
    form_contrato.sincronizar_contrato(fecha_ingreso)
    datos.save()
    return trabajador


@login_required
def trabajador_create(request):
    if not request.user.puede_editar:
        messages.error(request, "Tu rol no permite crear expedientes.")
        return redirect("expedientes:trabajador_list")

    form = TrabajadorForm(request.POST or None, usuario=request.user, creando=True)
    form_contrato = DatosContratacionForm(request.POST or None)

    if request.method == "POST":
        trabajador = _guardar_expediente(request, form, form_contrato)
        if trabajador is not None:
            registrar(request, RegistroAuditoria.Accion.CREAR, entidad="Trabajador",
                      objeto_id=trabajador.pk,
                      descripcion=f"Creó expediente de {trabajador}")
            messages.success(request, "Expediente creado correctamente.")
            return redirect("expedientes:trabajador_detail", pk=trabajador.pk)

    secciones = _secciones_expediente(form, form_contrato)
    return render(request, "expedientes/trabajador_form.html", {
        "form": form, "form_contrato": form_contrato, "titulo": "Nuevo expediente",
        "secciones": secciones,
        "problemas": _problemas_del_expediente(secciones, form, form_contrato),
    })


@login_required
def trabajador_update(request, pk):
    trabajador = get_object_or_404(Trabajador, pk=pk)
    exigir_editar_trabajador(request.user, trabajador)

    datos, _ = DatosContratacion.objects.get_or_create(trabajador=trabajador)
    form = TrabajadorForm(request.POST or None, instance=trabajador, usuario=request.user)
    form_contrato = DatosContratacionForm(request.POST or None, instance=datos)

    if request.method == "POST":
        guardado = _guardar_expediente(request, form, form_contrato, trabajador)
        if guardado is not None:
            registrar(request, RegistroAuditoria.Accion.EDITAR, entidad="Trabajador",
                      objeto_id=trabajador.pk,
                      descripcion=f"Editó datos de {trabajador}")
            messages.success(request, "Datos actualizados.")
            return redirect("expedientes:trabajador_detail", pk=trabajador.pk)

    secciones = _secciones_expediente(form, form_contrato)
    return render(request, "expedientes/trabajador_form.html", {
        "form": form, "form_contrato": form_contrato,
        "titulo": f"Editar: {trabajador.nombre_completo}",
        "secciones": secciones,
        "problemas": _problemas_del_expediente(secciones, form, form_contrato),
    })


@login_required
@require_POST
def trabajador_estado(request, pk):
    """Da de baja o reactiva el expediente. No se borra nada: solo cambia
    cómo cuenta en los listados (activo o baja)."""
    trabajador = get_object_or_404(Trabajador, pk=pk)
    exigir_editar_trabajador(request.user, trabajador)

    if trabajador.estado == Trabajador.Estado.ACTIVO:
        trabajador.estado = Trabajador.Estado.BAJA
        trabajador.observaciones_baja = request.POST.get("observaciones", "").strip()[:400]
        accion = "Dio de baja"
    else:
        trabajador.estado = Trabajador.Estado.ACTIVO
        # Se limpia: la ficha muestra la baja vigente, y la anterior ya quedó
        # asentada en la auditoría con su texto.
        trabajador.observaciones_baja = ""
        accion = "Reactivó"
    trabajador.save(update_fields=["estado", "observaciones_baja", "actualizado_en"])

    detalle = f"{accion} el expediente de {trabajador}"
    if trabajador.observaciones_baja:
        detalle += f". Observaciones: {trabajador.observaciones_baja}"
    registrar(request, RegistroAuditoria.Accion.EDITAR, entidad="Trabajador",
              objeto_id=trabajador.pk, descripcion=detalle)
    messages.success(request, f"{accion} el expediente de {trabajador}.")
    return redirect("expedientes:trabajador_detail", pk=trabajador.pk)


@login_required
@require_POST
def hijo_agregar(request, pk):
    """Agrega un hijo al expediente."""
    trabajador = get_object_or_404(Trabajador.objects.select_related("sede"), pk=pk)
    exigir_editar_trabajador(request.user, trabajador)

    form = HijoForm(request.POST, trabajador=trabajador)
    if form.is_valid():
        hijo = form.save(commit=False)
        hijo.trabajador = trabajador
        hijo.creado_por = request.user
        hijo.save()
        registrar(request, RegistroAuditoria.Accion.CREAR, entidad="Hijo",
                  objeto_id=hijo.pk,
                  descripcion=f"Agregó a {hijo.nombre_completo} como hijo/a de "
                              f"{trabajador}")
        messages.success(request, f"Agregado: {hijo.nombre_completo}.")
    else:
        messages.error(request, f"No se pudo agregar. {_errores_legibles(form)}")
    return redirect("expedientes:trabajador_detail", pk=trabajador.pk)


@login_required
@require_POST
def hijo_borrar(request, pk):
    """Saca un hijo del expediente. Se borra de verdad: es un dato de ficha.

    Solo el Administrador: los demás roles cargan y corrigen, pero no quitan.
    """
    hijo = get_object_or_404(
        Hijo.objects.select_related("trabajador", "trabajador__sede"), pk=pk
    )
    trabajador = hijo.trabajador
    exigir_borrar_del_expediente(request.user, trabajador)

    nombre = hijo.nombre_completo
    hijo.delete()
    registrar(request, RegistroAuditoria.Accion.BORRAR, entidad="Hijo",
              objeto_id=pk,
              descripcion=f"Quitó a {nombre} de los hijos de {trabajador}")
    messages.success(request, f"Se quitó a {nombre}.")
    return redirect("expedientes:trabajador_detail", pk=trabajador.pk)


def _proxima_version(trabajador, tipo):
    """Cuenta los documentos previos de ese tipo, incluidos los de la papelera.

    Contar solo los activos haría que al borrar uno el siguiente repitiera
    número, y dos versiones distintas con el mismo número no se distinguen.
    """
    return trabajador.documentos.filter(tipo=tipo).count() + 1


@login_required
@require_POST
def documento_subir(request, pk):
    trabajador = get_object_or_404(Trabajador, pk=pk)
    exigir_editar_trabajador(request.user, trabajador)

    form = DocumentoForm(request.POST, request.FILES)
    if form.is_valid():
        doc = form.save(commit=False)
        doc.trabajador = trabajador
        doc.subido_por = request.user
        doc.nombre_original = request.FILES["archivo"].name[:255]
        doc.tamano_bytes = request.FILES["archivo"].size
        doc.version = _proxima_version(trabajador, doc.tipo)
        doc.save()
        registrar(request, RegistroAuditoria.Accion.SUBIR, entidad="Documento",
                  objeto_id=doc.pk,
                  descripcion=f"Subió '{doc.tipo}' (v{doc.version}) a {trabajador}")
        messages.success(request, f"Documento '{doc.tipo}' cargado (v{doc.version}).")
    else:
        errores = "; ".join(f"{c}: {', '.join(e)}" for c, e in form.errors.items())
        messages.error(request, f"No se pudo subir el documento. {errores}")
    return redirect("expedientes:trabajador_detail", pk=trabajador.pk)


@login_required
@require_POST
def documento_comprimir(request, pk):
    """Recibe un documento que pasaba del tope, lo comprime y lo guarda.

    Es la mitad servidor del botón "Comprimir aquí y subir": las imágenes se
    comprimen en el navegador (ver `static/js/subida.js`), pero un PDF hay que
    redibujarlo página por página y eso se hace acá, con PyMuPDF.

    Responde JSON como el escáner: la pantalla no recarga. El archivo puede
    llegar hasta COMPRESION_MAX_BYTES (el middleware abre esa excepción solo
    para esta ruta), pero lo que se guarda queda por debajo de
    DOCUMENTOS_MAX_BYTES o se rechaza.
    """
    from django.core.files.base import ContentFile

    trabajador = get_object_or_404(Trabajador, pk=pk)
    exigir_editar_trabajador(request.user, trabajador)

    # EscaneoForm y no DocumentoForm: este archivo VIENE pasado del tope y el
    # validador de tamaño del modelo lo rechazaría antes de poder comprimirlo.
    form = EscaneoForm(request.POST)
    if not form.is_valid():
        errores = "; ".join(f"{c}: {', '.join(e)}" for c, e in form.errors.items())
        return JsonResponse({"ok": False, "error": errores}, status=400)

    subido = request.FILES.get("archivo")
    if not subido:
        return JsonResponse({"ok": False, "error": "No llegó ningún archivo."},
                            status=400)
    extension = os.path.splitext(subido.name)[1].lower()
    try:
        if extension == ".pdf":
            # Con archivos grandes, mejor la ruta del temporal que la RAM.
            ruta_temp = getattr(subido, "temporary_file_path", None)
            datos = comprimir_pdf(ruta_temp() if ruta_temp else subido.read())
        elif extension in {".jpg", ".jpeg", ".png", ".tiff", ".tif", ".webp"}:
            datos = comprimir_imagen(subido.read())
            extension = ".jpg"
        else:
            return JsonResponse(
                {"ok": False,
                 "error": f"Los archivos '{extension}' no se pueden comprimir. "
                          "Convertilo a PDF o a imagen y probá de nuevo."},
                status=400)
    except NoSePudoComprimir as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)

    tipo = form.cleaned_data["tipo"]
    version = _proxima_version(trabajador, tipo)
    doc = Documento(
        trabajador=trabajador, tipo=tipo, subido_por=request.user,
        fecha_vencimiento=form.cleaned_data.get("fecha_vencimiento"),
        observaciones=form.cleaned_data.get("observaciones", ""),
        version=version, tamano_bytes=len(datos),
    )
    doc.nombre_original = (os.path.splitext(subido.name)[0][:250] + extension)
    doc.archivo.save(doc.nombre_original, ContentFile(datos), save=False)
    doc.save()

    def _mb(n):
        return f"{n / 1024 / 1024:.1f}".replace(".", ",") + " MB"

    registrar(request, RegistroAuditoria.Accion.SUBIR, entidad="Documento",
              objeto_id=doc.pk,
              descripcion=f"Subió '{tipo}' (v{version}) a {trabajador} "
                          f"(comprimido de {_mb(subido.size)} a {_mb(len(datos))})")
    messages.success(
        request, f"Documento '{tipo}' comprimido y cargado (v{version}).")
    return JsonResponse({"ok": True, "documento": doc.pk})


@login_required
@require_POST
def documento_escanear(request, pk):
    """Recibe las fotos del escáner del teléfono y las guarda como un PDF.

    Responde JSON porque la pantalla del escáner no recarga: muestra el error
    ahí mismo, sin que la persona pierda las fotos que ya sacó.
    """
    from django.core.files.base import ContentFile

    from .escaner import EscaneoInvalido, armar_pdf

    trabajador = get_object_or_404(Trabajador, pk=pk)
    exigir_editar_trabajador(request.user, trabajador)

    form = EscaneoForm(request.POST)
    if not form.is_valid():
        errores = "; ".join(f"{c}: {', '.join(e)}" for c, e in form.errors.items())
        return JsonResponse({"ok": False, "error": errores}, status=400)

    try:
        pdf = armar_pdf(request.FILES.getlist("paginas"))
    except EscaneoInvalido as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)

    tipo = form.cleaned_data["tipo"]
    version = _proxima_version(trabajador, tipo)
    doc = Documento(
        trabajador=trabajador, tipo=tipo, subido_por=request.user,
        fecha_vencimiento=form.cleaned_data.get("fecha_vencimiento"),
        observaciones=form.cleaned_data.get("observaciones", ""),
        version=version, tamano_bytes=len(pdf),
    )
    hojas = len(request.FILES.getlist("paginas"))
    # El nombre lo pone quien escanea. Si no puso nada, uno que al menos diga
    # de qué se trata: "escaneo-20260819-2h.pdf" no ayuda a nadie a encontrarlo.
    elegido = form.cleaned_data.get("nombre") or f"{tipo} {timezone.localdate():%d-%m-%Y}"
    doc.nombre_original = f"{elegido}.pdf"
    doc.archivo.save(doc.nombre_original, ContentFile(pdf), save=False)
    doc.save()

    registrar(request, RegistroAuditoria.Accion.SUBIR, entidad="Documento",
              objeto_id=doc.pk,
              descripcion=f"Escaneó '{tipo}' (v{version}, {hojas} hoja"
                          f"{'s' if hojas != 1 else ''}) de {trabajador}")
    messages.success(
        request,
        f"Documento '{tipo}' escaneado (v{version}, {hojas} hoja"
        f"{'s' if hojas != 1 else ''}).")
    return JsonResponse({"ok": True, "documento": doc.pk, "hojas": hojas})


@login_required
def documento_descargar(request, pk):
    doc = get_object_or_404(
        Documento.objects.select_related("trabajador", "trabajador__sede"), pk=pk
    )
    exigir_ver_trabajador(request.user, doc.trabajador)
    if not doc.activo and not request.user.es_admin:
        raise Http404()

    almacen = doc.archivo.storage
    if not almacen.exists(doc.archivo.name):
        raise Http404("El archivo no está disponible.")

    registrar(request, RegistroAuditoria.Accion.DESCARGAR, entidad="Documento",
              objeto_id=doc.pk,
              descripcion=f"Descargó '{doc.tipo}' de {doc.trabajador}")

    nombre = doc.nombre_original or os.path.basename(doc.archivo.name)
    tipo_mime = mimetypes.guess_type(nombre)[0] or "application/octet-stream"
    inline = doc.extension in {".pdf", ".jpg", ".jpeg", ".png", ".webp"}

    # Sale de a pedazos, descifrando sobre la marcha. Juntarlo entero en memoria
    # ponía el peso del archivo —por cada persona que descargue a la vez— en la
    # RAM del servidor, que es la mitad del problema que tenía subir.
    respuesta = FileResponse(almacen.pedazos_descifrados(doc.archivo.name),
                             content_type=tipo_mime)
    # Armado por Django y no a mano: ahora el nombre lo escribe una persona y
    # trae acentos, comillas y espacios. Pegado crudo entre comillas, "Cédula
    # de José" se descargaba con el nombre roto.
    respuesta["Content-Disposition"] = content_disposition_header(
        as_attachment=not inline, filename=nombre)
    return respuesta


def _volver(request, por_defecto, *args):
    """A dónde volver después de una acción sobre un documento.

    La misma acción se dispara desde dos lados —el expediente y la lista de
    pendientes de Configuración— y cada uno tiene que volver a lo suyo. El
    destino llega en el formulario, así que se valida igual que el `next` del
    login: solo se acepta una dirección de este mismo sistema.
    """
    destino = request.POST.get("volver")
    if destino and url_has_allowed_host_and_scheme(
            url=destino, allowed_hosts={request.get_host()},
            require_https=request.is_secure()):
        return redirect(destino)
    return redirect(por_defecto, *args)


@login_required
@require_POST
def documento_marcar(request, pk):
    """Pedir que se elimine un documento. No lo borra: lo deja en la lista.

    Lo puede hacer quien puede subir. Es la mitad que faltaba: hasta ahora
    quien se daba cuenta de que subió el archivo equivocado no tenía forma de
    decirlo, y borrar sigue siendo del Administrador.
    """
    doc = get_object_or_404(Documento.objects.select_related("trabajador"), pk=pk)
    exigir_editar_trabajador(request.user, doc.trabajador)

    doc.marcar(request.user, request.POST.get("motivo", ""))
    registrar(request, RegistroAuditoria.Accion.EDITAR, entidad="Documento",
              objeto_id=doc.pk,
              descripcion=f"Marcó para eliminar '{doc.tipo}' de {doc.trabajador}"
                          + (f": {doc.motivo_marca}" if doc.motivo_marca else ""))
    messages.warning(
        request,
        "Documento marcado para eliminar. Queda en el expediente hasta que el "
        "Administrador lo confirme desde Configuración.")
    return _volver(request, "expedientes:trabajador_detail", doc.trabajador_id)


@login_required
@require_POST
def documento_desmarcar(request, pk):
    """El «en caso tal»: sacar la marca."""
    doc = get_object_or_404(Documento.objects.select_related("trabajador"), pk=pk)
    exigir_editar_trabajador(request.user, doc.trabajador)

    doc.desmarcar()
    registrar(request, RegistroAuditoria.Accion.EDITAR, entidad="Documento",
              objeto_id=doc.pk,
              descripcion=f"Quitó la marca de eliminar de '{doc.tipo}' "
                          f"de {doc.trabajador}")
    messages.success(request, "Se quitó la marca: el documento se queda.")
    return _volver(request, "expedientes:trabajador_detail", doc.trabajador_id)


@login_required
@require_POST
def documento_borrar(request, pk):
    doc = get_object_or_404(Documento.objects.select_related("trabajador"), pk=pk)
    exigir_ver_trabajador(request.user, doc.trabajador)
    exigir_borrar(request.user)

    doc.activo = False
    doc.save(update_fields=["activo"])
    registrar(request, RegistroAuditoria.Accion.BORRAR, entidad="Documento",
              objeto_id=doc.pk,
              descripcion=f"Envió a papelera '{doc.tipo}' de {doc.trabajador}")
    messages.warning(request, "Documento enviado a la papelera.")
    return _volver(request, "expedientes:trabajador_detail", doc.trabajador_id)


@login_required
@require_POST
def documento_restaurar(request, pk):
    doc = get_object_or_404(Documento.objects.select_related("trabajador"), pk=pk)
    exigir_ver_trabajador(request.user, doc.trabajador)
    exigir_borrar(request.user)

    doc.activo = True
    doc.save(update_fields=["activo"])
    registrar(request, RegistroAuditoria.Accion.RESTAURAR, entidad="Documento",
              objeto_id=doc.pk,
              descripcion=f"Restauró '{doc.tipo}' de {doc.trabajador}")
    messages.success(request, "Documento restaurado.")
    return redirect("expedientes:trabajador_detail", pk=doc.trabajador_id)


@login_required
def papelera(request, pk):
    """Documentos en papelera de un trabajador (solo admin)."""
    trabajador = get_object_or_404(Trabajador, pk=pk)
    exigir_ver_trabajador(request.user, trabajador)
    if not request.user.es_admin:
        messages.error(request, "Solo el Administrador puede ver la papelera.")
        return redirect("expedientes:trabajador_detail", pk=pk)

    documentos = trabajador.documentos.filter(activo=False).select_related("tipo")
    return render(request, "expedientes/papelera.html",
                  {"trabajador": trabajador, "documentos": documentos})


def _nomina_filtrada(request):
    """Queryset de trabajadores visibles + filtros de la nómina. Devuelve (form, qs)."""
    form = FiltroTrabajadorForm(request.GET or None, usuario=request.user)
    qs = (trabajadores_visibles(request.user)
          .select_related("sede", "sede__zona", "departamento", "contratacion",
                          "puesto", "tipo_documento")
          .prefetch_related("hijos")
          .order_by("apellidos", "nombres"))
    # Se aplican los filtros que hayan validado, uno por uno. Con
    # `if form.is_valid()` un solo valor raro en la URL —una tienda borrada, un
    # estado inventado— tiraba abajo TODOS los filtros y devolvía la nómina
    # entera, que es justo lo contrario de lo que espera quien filtró.
    form.is_valid()
    datos = getattr(form, "cleaned_data", {})
    if datos:
        q = datos.get("q")
        sedes = datos.get("sedes")
        departamento = datos.get("departamento")
        estado = datos.get("estado")
        if q:
            qs = qs.filter(
                Q(nombres__icontains=q)
                | Q(apellidos__icontains=q)
                | Q(documento_identidad__icontains=q)
            )
        if sedes:
            qs = qs.filter(sede__in=sedes)
        if departamento:
            qs = qs.filter(departamento=departamento)
        if estado:
            qs = qs.filter(estado=estado)
    return form, qs


@login_required
def nomina(request):
    """Listado de trabajadores: C.I., apellidos, nombres, cargo, departamento, tienda."""
    form, qs = _nomina_filtrada(request)
    paginador = Paginator(qs, 25)
    pagina = paginador.get_page(request.GET.get("page"))

    # Semáforo de datos: cuánto de lo que va al Excel está cargado. Se calcula
    # sobre los objetos que ya trajo la página —`select_related` de contratación
    # incluido—, así que no agrega ni una consulta.
    con_pagos = request.user.puede_editar
    for t in pagina:
        t.completitud = completitud(t, con_pagos)

    parcial = bool(request.headers.get("HX-Request"))
    contexto = {"form": form, "pagina": pagina, "total": qs.count(),
                "querystring": _querystring_sin_pagina(request),
                "parcial": parcial}
    if parcial:
        # Solo se reemplaza la tabla, así que el conteo del encabezado viaja
        # aparte (hx-swap-oob). Si no, quedaría con el número de antes de filtrar.
        return render(request, "expedientes/_tabla_nomina.html", contexto)
    return render(request, "expedientes/nomina.html", contexto)


def _conceptos_para_export(trabajadores):
    """Columnas de pago del Excel: una por concepto, con el monto de cada quien.

    Devuelve (conceptos, montos):
      - conceptos: los ConceptoPago que llevan columna, en el orden del catálogo.
      - montos[trabajador_id][concepto_id] = monto de esa persona en ese concepto.

    No se totaliza nada: cada concepto tiene su moneda y no hay tasa de cambio,
    así que sumarlos entre sí no significaría nada.

    Llevan columna todos los conceptos activos —aunque hoy no los cobre nadie,
    para que el formato del archivo no cambie de un mes a otro— más los que
    estén inactivos pero sigan asignados, para no perder plata ya registrada.
    """
    pagos = (
        AsignacionPago.objects
        .filter(activo=True, trabajador__in=trabajadores, concepto__isnull=False)
        .select_related("concepto", "concepto__moneda")
    )

    conceptos = {c.pk: c for c in
                 ConceptoPago.objects.filter(activo=True).select_related("moneda")}
    montos = {}
    for p in pagos:
        conceptos.setdefault(p.concepto_id, p.concepto)
        fila = montos.setdefault(p.trabajador_id, {})
        # Sumar es por si quedó una asignación repetida de una carga vieja:
        # la grilla del expediente ya deja una sola por concepto.
        fila[p.concepto_id] = fila.get(p.concepto_id, Decimal("0")) + p.monto

    ordenados = sorted(conceptos.values(), key=lambda c: (c.orden, c.nombre))
    return ordenados, montos


# Ancho del bloque final del Excel, en celdas. Son 7 porque así se recorta y se
# pega en los formatos en papel que ya se usan en RRHH.
CELDAS_DEL_BLOQUE = 7


def _bloque_nombre_y_cedula(ws, trabajadores, ultima_columna):
    """Agrega al final una celda ancha con el nombre y la cédula juntos.

    Es para recortar y pegar: una sola celda por persona, en negrita y
    centrada, con el texto achicándose solo si no entra (`shrink_to_fit`) en vez
    de desbordarse sobre la celda de al lado o quedar cortado.

    Las columnas de la izquierda siguen teniendo el nombre y la cédula por
    separado: esto se suma, no las reemplaza, para que el archivo siga
    sirviendo para filtrar y ordenar.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    primera = ultima_columna + 1
    ultima = ultima_columna + CELDAS_DEL_BLOQUE
    linea = Side(style="thin", color="000000")
    recuadro = Border(left=linea, right=linea, top=linea, bottom=linea)

    # Encabezado, con el mismo rojo que el resto.
    ws.merge_cells(start_row=1, start_column=primera, end_row=1, end_column=ultima)
    titulo = ws.cell(row=1, column=primera, value="Nombre y cédula")
    titulo.fill = PatternFill("solid", fgColor="E1052D")
    titulo.font = Font(bold=True, color="FFFFFF")
    titulo.alignment = Alignment(horizontal="center", vertical="center")

    for indice, t in enumerate(trabajadores):
        fila = indice + 2
        ws.merge_cells(start_row=fila, start_column=primera,
                       end_row=fila, end_column=ultima)
        celda = ws.cell(row=fila, column=primera,
                        value=f"{t.apellidos} {t.nombres}   {t.documento_identidad}")
        celda.font = Font(bold=True, size=12)
        # shrink_to_fit es lo que pidió el usuario con "ajustable al tamaño de
        # las casillas": si achica la columna, Excel achica la letra.
        celda.alignment = Alignment(horizontal="center", vertical="center",
                                    shrink_to_fit=True)
        # El recuadro se pinta celda por celda: en un rango combinado, el borde
        # de la primera no dibuja la caja entera.
        for col in range(primera, ultima + 1):
            ws.cell(row=fila, column=col).border = recuadro

    for col in range(primera, ultima + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9


@login_required
def nomina_export(request):
    """Exporta la nómina filtrada a un archivo Excel (.xlsx).

    Cada concepto del catálogo (Configuración → Conceptos de pago) es una
    columna con su moneda en el encabezado. No hay totales: las monedas no se
    suman entre sí porque el sistema no guarda tasas de cambio.

    Los datos de pago —cuenta bancaria y montos— solo salen si el usuario puede
    verlos (Administrador o RRHH Interior). Solo lectura recibe el mismo listado
    sin esas columnas: con qué banco y cuánto cobra cada persona es justamente
    lo que ese rol no tiene por qué ver.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _, qs = _nomina_filtrada(request)
    trabajadores = list(qs)

    incluir_pagos = request.user.puede_editar
    conceptos, montos = (_conceptos_para_export(trabajadores) if incluir_pagos
                         else ([], {}))

    # (encabezado, cómo sacar el valor, formato de la celda). Armar las columnas
    # así evita que los índices de formato se desfasen al agregar una columna.
    TEXTO, FECHA, MONTO = "@", "DD/MM/YYYY", "#,##0.00"

    def contratacion(t):
        return getattr(t, "contratacion", None)

    def de_contratacion(campo):
        def leer(t):
            datos = contratacion(t)
            return getattr(datos, campo) if datos else ""
        return leer

    columnas = [
        ("C.I.", lambda t: t.cedula_completa, TEXTO),
        ("Apellidos", lambda t: t.apellidos, None),
        ("Nombres", lambda t: t.nombres, None),
        ("Cargo", lambda t: t.cargo_nombre, None),
        ("Departamento", lambda t: t.departamento.nombre if t.departamento else "", None),
        ("Tienda", lambda t: t.sede.nombre if t.sede else "", None),
        ("Fecha de ingreso", lambda t: t.fecha_ingreso, FECHA),
    ]

    if incluir_pagos:
        # Los números de cuenta van como TEXTO a propósito: si Excel los tomara
        # como número se comería el cero inicial (0102… -> 102…) y la
        # transferencia iría a una cuenta inexistente.
        columnas += [
            ("Banco", de_contratacion("banco"), None),
            ("Prefijo", de_contratacion("prefijo"), TEXTO),
            ("Número de cuenta", de_contratacion("numero_cuenta"), TEXTO),
            ("Cuenta bancaria", de_contratacion("cuenta_bancaria"), TEXTO),
        ]

    columnas += [
        ("Hijos", lambda t: t.cantidad_hijos, None),
        ("Talla camisa", de_contratacion("talla_camisa"), None),
        ("Talla pantalón", de_contratacion("talla_pantalon"), None),
        ("Talla zapato", de_contratacion("talla_zapato"), None),
    ]

    # El encabezado dice el concepto tal cual se llama en el catálogo, con su
    # moneda, para que en la celda quede solo el número y Excel pueda operarlo.
    for c in conceptos:
        columnas.append((
            f"{c.nombre} ({c.moneda.simbolo})",
            (lambda concepto: lambda t: montos.get(t.pk, {}).get(concepto.pk))(c),
            MONTO,
        ))

    wb = Workbook()
    ws = wb.active
    ws.title = "Nómina"
    ws.append([titulo for titulo, _, _ in columnas])

    # Estilo del encabezado (rojo Damasco, texto blanco).
    header_fill = PatternFill("solid", fgColor="E1052D")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(columnas) + 1):
        celda = ws.cell(row=1, column=col)
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal="left", vertical="center")

    for t in trabajadores:
        # Un dato que no está -> celda vacía, no un 0 ni un "None" que confundan.
        ws.append([valor(t) for _, valor, _ in columnas])

    for col, (_titulo, _valor, formato) in enumerate(columnas, start=1):
        if not formato:
            continue
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = formato

    # Ancho de columnas automático (acotado).
    for col in range(1, len(columnas) + 1):
        letra = get_column_letter(col)
        largo = max([len(str(columnas[col - 1][0]))]
                    + [len(str(ws.cell(row=r, column=col).value or ""))
                       for r in range(2, ws.max_row + 1)] or [0])
        ws.column_dimensions[letra].width = min(max(largo + 2, 12), 40)

    _bloque_nombre_y_cedula(ws, trabajadores, len(columnas))
    ws.freeze_panes = "A2"

    registrar(request, RegistroAuditoria.Accion.DESCARGAR, entidad="Nómina",
              descripcion=f"Exportó la nómina a Excel ({len(trabajadores)} trabajadores"
                          f"{', con datos bancarios y montos' if incluir_pagos else ''})")

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fecha = timezone.localdate().strftime("%Y%m%d")
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="nomina_{fecha}.xlsx"'
    return resp


@login_required
def auditoria_list(request):
    """Bitácora de auditoría. Solo administradores."""
    if not request.user.es_admin:
        messages.error(request, "Solo el Administrador puede ver la auditoría.")
        return redirect("expedientes:trabajador_list")

    qs = RegistroAuditoria.objects.select_related("usuario")
    accion = request.GET.get("accion")
    if accion:
        qs = qs.filter(accion=accion)
    q = request.GET.get("q")
    if q:
        qs = qs.filter(Q(usuario_texto__icontains=q) | Q(descripcion__icontains=q))

    paginador = Paginator(qs, 30)
    pagina = paginador.get_page(request.GET.get("page"))
    return render(request, "expedientes/auditoria_list.html", {
        "pagina": pagina,
        "acciones": RegistroAuditoria.Accion.choices,
        "accion_sel": accion or "",
        "q": q or "",
    })
