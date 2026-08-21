"""El bloque final del Excel: nombre y cédula juntos, para recortar y pegar.

Una celda por persona, ancha (siete casillas combinadas), en negrita, centrada
y con el texto que se achica solo si no entra. Las columnas sueltas de nombre y
cédula siguen estando: esto se suma, no las reemplaza.
"""

from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from cuentas.models import Sede, Zona
from expedientes.models import Trabajador
from expedientes.views import CELDAS_DEL_BLOQUE

Usuario = get_user_model()
CLAVE = "Clave-De-Prueba-123"


class BloqueNombreYCedula(TestCase):

    @classmethod
    def setUpTestData(cls):
        zona = Zona.objects.create(nombre="ZULIA")
        cls.sede = Sede.objects.create(nombre="MARACAIBO", zona=zona)
        cls.andres = Trabajador.objects.create(
            documento_identidad="26276490", nombres="ANDRES DAVID",
            apellidos="PARRA MARIN", sede=cls.sede)
        cls.gerson = Trabajador.objects.create(
            documento_identidad="27886008", nombres="GERSON ALEJANDRO",
            apellidos="SANCHEZ JAIMES", sede=cls.sede)

        cls.admin = Usuario.objects.create_user(username="adm", password=CLAVE)
        cls.admin.rol = Usuario.Rol.ADMIN
        cls.admin.save()

        cls.lectura = Usuario.objects.create_user(username="lect", password=CLAVE)
        cls.lectura.rol = Usuario.Rol.SOLO_LECTURA
        cls.lectura.save()

    def hoja(self, usuario=None, **filtros):
        self.client.force_login(usuario or self.admin)
        r = self.client.get(reverse("expedientes:nomina_export"), filtros)
        self.assertEqual(r.status_code, 200)
        return load_workbook(BytesIO(r.content)).active

    def primera_del_bloque(self, hoja):
        """Columna donde arranca el bloque: la que sigue al encabezado 'Talla zapato'…

        …o a la última que tenga título. Se busca, en vez de fijar un número,
        para que agregar una columna a la nómina no rompa este test.
        """
        for col in range(1, hoja.max_column + 1):
            if hoja.cell(row=1, column=col).value == "Nombre y cédula":
                return col
        raise AssertionError("no está el bloque de nombre y cédula")

    # --- Contenido ------------------------------------------------------------
    def test_esta_al_final_de_todo(self):
        hoja = self.hoja()
        primera = self.primera_del_bloque(hoja)
        self.assertEqual(primera + CELDAS_DEL_BLOQUE - 1, hoja.max_column)

    def test_trae_el_nombre_y_la_cedula_juntos(self):
        hoja = self.hoja()
        col = self.primera_del_bloque(hoja)
        textos = [hoja.cell(row=r, column=col).value
                  for r in range(2, hoja.max_row + 1)]
        self.assertIn("PARRA MARIN ANDRES DAVID   26276490", textos)
        self.assertIn("SANCHEZ JAIMES GERSON ALEJANDRO   27886008", textos)

    def test_la_cedula_no_se_convierte_en_numero(self):
        """Con separadores de miles ya no sirve para copiar a un formulario."""
        hoja = self.hoja()
        col = self.primera_del_bloque(hoja)
        valor = hoja.cell(row=2, column=col).value
        self.assertIsInstance(valor, str)
        self.assertNotIn(",", valor)
        self.assertNotIn(".", valor)

    def test_las_columnas_sueltas_siguen_estando(self):
        hoja = self.hoja()
        encabezados = [c.value for c in hoja[1]]
        for titulo in ("C.I.", "Apellidos", "Nombres"):
            self.assertIn(titulo, encabezados)

    def test_hay_una_fila_por_persona_y_ninguna_de_más(self):
        hoja = self.hoja()
        col = self.primera_del_bloque(hoja)
        llenas = [r for r in range(2, hoja.max_row + 1)
                  if hoja.cell(row=r, column=col).value]
        self.assertEqual(len(llenas), 2)

    # --- Formato --------------------------------------------------------------
    def test_son_siete_casillas_unidas(self):
        hoja = self.hoja()
        col = self.primera_del_bloque(hoja)
        rangos = {str(r) for r in hoja.merged_cells.ranges}
        letra_a = get_column_letter(col)
        letra_b = get_column_letter(col + CELDAS_DEL_BLOQUE - 1)
        self.assertIn(f"{letra_a}2:{letra_b}2", rangos)
        self.assertIn(f"{letra_a}3:{letra_b}3", rangos)

    def test_la_letra_va_en_negrita(self):
        hoja = self.hoja()
        celda = hoja.cell(row=2, column=self.primera_del_bloque(hoja))
        self.assertTrue(celda.font.bold)

    def test_el_texto_se_achica_para_entrar(self):
        """"Ajustable al tamaño de las casillas": si achican la columna, achica la letra."""
        hoja = self.hoja()
        celda = hoja.cell(row=2, column=self.primera_del_bloque(hoja))
        self.assertTrue(celda.alignment.shrink_to_fit)
        self.assertEqual(celda.alignment.horizontal, "center")

    def test_cada_celda_del_bloque_lleva_recuadro(self):
        """El borde se pinta en las siete: solo en la primera no dibuja la caja."""
        hoja = self.hoja()
        col = self.primera_del_bloque(hoja)
        for c in range(col, col + CELDAS_DEL_BLOQUE):
            celda = hoja.cell(row=2, column=c)
            self.assertEqual(celda.border.top.style, "thin", f"columna {c}")
            self.assertEqual(celda.border.bottom.style, "thin", f"columna {c}")

    def test_el_encabezado_del_bloque_va_combinado_y_en_rojo(self):
        hoja = self.hoja()
        col = self.primera_del_bloque(hoja)
        titulo = hoja.cell(row=1, column=col)
        self.assertEqual(titulo.value, "Nombre y cédula")
        self.assertTrue(titulo.font.bold)
        self.assertIn("E1052D", titulo.fill.fgColor.rgb or "")

    # --- Se lleva bien con el resto -------------------------------------------
    def test_respeta_el_filtro(self):
        hoja = self.hoja(q="SANCHEZ")
        col = self.primera_del_bloque(hoja)
        textos = [hoja.cell(row=r, column=col).value
                  for r in range(2, hoja.max_row + 1)]
        self.assertEqual(textos, ["SANCHEZ JAIMES GERSON ALEJANDRO   27886008"])

    def test_solo_lectura_tambien_lo_recibe(self):
        """No es dato de pago: no hay motivo para escondérselo."""
        hoja = self.hoja(self.lectura)
        col = self.primera_del_bloque(hoja)
        self.assertIn("PARRA MARIN", hoja.cell(row=2, column=col).value)
        self.assertNotIn("Cuenta bancaria", [c.value for c in hoja[1]])

    def test_sin_nadie_queda_solo_el_encabezado(self):
        Trabajador.objects.all().delete()
        hoja = self.hoja()
        col = self.primera_del_bloque(hoja)
        self.assertEqual(hoja.cell(row=1, column=col).value, "Nombre y cédula")
        self.assertIsNone(hoja.cell(row=2, column=col).value)
