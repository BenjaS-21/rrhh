"""El tipo de cédula: la letra que va delante del número.

Es un catálogo —V, E, J, P, G— que se administra desde el admin de Django, y un
campo pegado a la cédula en el alta y en la edición.

Lo que más importa acá es lo de siempre con un dato nuevo: que no rompa los
expedientes que ya estaban. Los viejos traen la letra metida dentro del número
(«V-30719983») y no tienen tipo; para esos no se antepone nada, o saldría
«V-V-30719983» en los contratos.
"""

import re

from django.contrib.auth import get_user_model
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import reverse

from cuentas.models import Sede, TipoDocumentoIdentidad, Zona
from expedientes.models import Trabajador

Usuario = get_user_model()
CLAVE = "Clave-De-Prueba-123"


class ElCatalogoVieneCargado(TestCase):
    """Un catálogo vacío obligaría a entrar al admin antes del primer alta."""

    def test_estan_los_tipos_venezolanos(self):
        codigos = set(TipoDocumentoIdentidad.objects.values_list("codigo", flat=True))
        self.assertEqual(codigos, {"V", "E", "J", "P", "G"})

    def test_vienen_activos_y_ordenados(self):
        tipos = list(TipoDocumentoIdentidad.objects.filter(activo=True))
        self.assertEqual([t.codigo for t in tipos], ["V", "E", "J", "P", "G"])

    def test_el_codigo_se_guarda_en_mayusculas_y_sin_puntos(self):
        """Es lo que se imprime en los documentos: «v.» y «V» no pueden convivir."""
        t = TipoDocumentoIdentidad.objects.create(codigo=" r. ", nombre="Prueba")
        self.assertEqual(t.codigo, "R")

    def test_se_administra_desde_el_admin_de_django(self):
        from django.contrib import admin
        self.assertIn(TipoDocumentoIdentidad, admin.site._registry)

    def test_y_no_desde_configuracion(self):
        """Testigo: se pidió expresamente que fuera solo en el admin."""
        from configuracion.views import CATALOGOS
        modelos = {c["model"] for c in CATALOGOS.values()}
        self.assertNotIn(TipoDocumentoIdentidad, modelos)


class _Base(TestCase):

    @classmethod
    def setUpTestData(cls):
        zona = Zona.objects.create(nombre="TACHIRA")
        cls.sede = Sede.objects.create(nombre="SAN CRISTOBAL", zona=zona)
        cls.v = TipoDocumentoIdentidad.objects.get(codigo="V")
        cls.e = TipoDocumentoIdentidad.objects.get(codigo="E")
        cls.admin = Usuario.objects.create_user(username="adm", password=CLAVE)
        cls.admin.rol = Usuario.Rol.ADMIN
        cls.admin.save()

    def alta(self, **cambios):
        datos = {"documento_identidad": "30719983", "nombres": "Benjamin",
                 "apellidos": "Velazco", "sede": self.sede.pk,
                 "tipo_documento": self.v.pk}
        datos.update(cambios)
        self.client.force_login(self.admin)
        return self.client.post(reverse("expedientes:trabajador_create"), datos)


class EnElAltaYLaEdicion(_Base):

    def formulario(self, url):
        self.client.force_login(self.admin)
        return self.client.get(url).content.decode()

    def test_el_campo_esta_en_el_alta(self):
        cuerpo = self.formulario(reverse("expedientes:trabajador_create"))
        self.assertIn('name="tipo_documento"', cuerpo)

    def test_y_en_la_edicion(self):
        t = Trabajador.objects.create(
            documento_identidad="1", nombres="Ana", apellidos="Alvarez",
            sede=self.sede)
        cuerpo = self.formulario(
            reverse("expedientes:trabajador_update", args=[t.pk]))
        self.assertIn('name="tipo_documento"', cuerpo)

    def test_esta_pegado_a_la_cedula_y_no_en_otra_parte(self):
        """Se pidió al lado de la cédula: comparten una sola casilla."""
        cuerpo = self.formulario(reverse("expedientes:trabajador_create"))
        casilla = re.search(r'<div class="campo campo--pareja">.*?</div>\s*</div>',
                            cuerpo, re.S)
        self.assertIsNotNone(casilla, "el tipo no quedó en la casilla de la cédula")
        self.assertIn('name="tipo_documento"', casilla.group(0))
        self.assertIn('name="documento_identidad"', casilla.group(0))

    def test_el_tipo_va_primero(self):
        """La letra se lee antes que el número, como está escrita en la cédula."""
        cuerpo = self.formulario(reverse("expedientes:trabajador_create"))
        self.assertLess(cuerpo.index('name="tipo_documento"'),
                        cuerpo.index('name="documento_identidad"'))

    def test_ofrece_los_cinco_tipos(self):
        cuerpo = self.formulario(reverse("expedientes:trabajador_create"))
        trozo = re.search(r'<select[^>]*name="tipo_documento".*?</select>',
                          cuerpo, re.S).group(0)
        self.assertEqual(len(re.findall(r"<option", trozo)), 6)   # 5 + el vacío

    def test_un_tipo_dado_de_baja_no_se_ofrece(self):
        TipoDocumentoIdentidad.objects.filter(codigo="G").update(activo=False)
        cuerpo = self.formulario(reverse("expedientes:trabajador_create"))
        trozo = re.search(r'<select[^>]*name="tipo_documento".*?</select>',
                          cuerpo, re.S).group(0)
        self.assertNotIn("Gubernamental", trozo)

    def test_pero_el_que_ya_tiene_el_expediente_sigue_estando(self):
        """Si no, abrir una ficha vieja y guardarla le borraría el tipo."""
        g = TipoDocumentoIdentidad.objects.get(codigo="G")
        t = Trabajador.objects.create(
            documento_identidad="2", nombres="Ana", apellidos="Alvarez",
            sede=self.sede, tipo_documento=g)
        TipoDocumentoIdentidad.objects.filter(codigo="G").update(activo=False)
        cuerpo = self.formulario(
            reverse("expedientes:trabajador_update", args=[t.pk]))
        trozo = re.search(r'<select[^>]*name="tipo_documento".*?</select>',
                          cuerpo, re.S).group(0)
        self.assertIn("Gubernamental", trozo)


class SeGuardaYSeLee(_Base):

    def test_se_guarda_lo_elegido(self):
        self.assertEqual(self.alta().status_code, 302)
        t = Trabajador.objects.get()
        self.assertEqual(t.tipo_documento, self.v)
        self.assertEqual(t.documento_identidad, "30719983")

    def test_la_cedula_completa_junta_los_dos(self):
        self.alta(tipo_documento=self.e.pk)
        self.assertEqual(Trabajador.objects.get().cedula_completa, "E-30719983")

    def test_no_es_obligatorio(self):
        """Los expedientes viejos no lo tienen: exigirlo trabaría toda edición."""
        self.assertEqual(self.alta(tipo_documento="").status_code, 302)
        self.assertIsNone(Trabajador.objects.get().tipo_documento)

    def test_sin_tipo_la_cedula_sale_tal_cual(self):
        """El testigo del arrastre: «V-30719983» no puede volverse «V-V-…»."""
        t = Trabajador.objects.create(
            documento_identidad="V-30719983", nombres="Ana", apellidos="Alvarez",
            sede=self.sede)
        self.assertEqual(t.cedula_completa, "V-30719983")

    def test_dos_personas_no_pueden_repetir_el_numero(self):
        """La cédula sigue siendo única: el tipo no abre una puerta de atrás."""
        self.alta()
        r = self.alta(nombres="Otro", apellidos="Distinto",
                      tipo_documento=self.e.pk)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(Trabajador.objects.count(), 1)

    def test_no_se_puede_borrar_un_tipo_que_esta_en_uso(self):
        """Borrarlo dejaría expedientes con media cédula, sin aviso."""
        from django.db.models import ProtectedError
        self.alta()
        with self.assertRaises(ProtectedError):
            self.v.delete()


class DondeSeMuestra(_Base):

    def setUp(self):
        self.alta()
        self.trabajador = Trabajador.objects.get()
        self.client.force_login(self.admin)

    def test_en_la_ficha(self):
        cuerpo = self.client.get(
            reverse("expedientes:trabajador_detail",
                    args=[self.trabajador.pk])).content.decode()
        self.assertIn("V-30719983", cuerpo)

    def test_en_el_listado_de_expedientes(self):
        cuerpo = self.client.get(
            reverse("expedientes:trabajador_list")).content.decode()
        self.assertIn("V-30719983", cuerpo)

    def test_en_la_nomina(self):
        cuerpo = self.client.get(reverse("expedientes:nomina")).content.decode()
        self.assertIn("V-30719983", cuerpo)

    def test_en_el_excel(self):
        from io import BytesIO
        from openpyxl import load_workbook
        r = self.client.get(reverse("expedientes:nomina_export"))
        hoja = load_workbook(BytesIO(r.content)).active
        valores = [c.value for fila in hoja.iter_rows() for c in fila]
        self.assertIn("V-30719983", valores)

    def test_en_los_documentos_word(self):
        from expedientes.documentos import contexto_documentos
        self.assertEqual(contexto_documentos(self.trabajador)["Cedula"],
                         "V-30719983")

    def test_buscar_por_el_numero_lo_encuentra(self):
        """Testigo: la gente busca por el número, con letra o sin ella."""
        r = self.client.get(reverse("expedientes:trabajador_list") + "?q=30719983")
        self.assertIn("VELAZCO", r.content.decode())


class SinConsultasDeMas(_Base):

    def _consultas(self, cuantos):
        """Cuántas consultas cuesta el listado con `cuantos` expedientes."""
        Trabajador.objects.all().delete()
        for i in range(cuantos):
            Trabajador.objects.create(
                documento_identidad=str(1000 + i), nombres=f"P{i}",
                apellidos="Prueba", sede=self.sede, tipo_documento=self.v)
        self.client.force_login(self.admin)
        with CaptureQueriesContext(connection) as capturadas:
            self.client.get(reverse("expedientes:trabajador_list"))
        return len(capturadas)

    def test_el_listado_no_pregunta_el_tipo_fila_por_fila(self):
        """No se fija un número mágico: se compara cómo escala.

        Sin `select_related`, cada fila del listado agrega una consulta para
        traer su tipo de cédula. Con 2 y con 20 personas tiene que costar lo
        mismo; si no, son 25 consultas de regalo por página.
        """
        pocas = self._consultas(2)
        muchas = self._consultas(20)
        self.assertEqual(pocas, muchas,
                         f"{pocas} consultas con 2 personas y {muchas} con 20: "
                         "el listado pregunta fila por fila")
