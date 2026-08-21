"""Tests de los datos bancarios en el Excel de nómina.

Lo delicado acá es el cero inicial: los prefijos bancarios venezolanos empiezan
con cero (0102, 0105…) y si Excel los toma como número se lo come, dejando una
cuenta que no existe. Por eso esas columnas van con formato de texto.
"""

from io import BytesIO

from django.urls import reverse
from openpyxl import load_workbook

from expedientes.models import DatosContratacion
from expedientes.tests import CLAVE, BasePagos


class DatosBancariosEnElExcel(BasePagos):

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        DatosContratacion.objects.create(
            trabajador=cls.trab_norte,
            banco="Banco de Venezuela",
            prefijo="0102",
            numero_cuenta="1234567890123456",
        )
        # Beto no tiene datos de contratación cargados: sus celdas van vacías.

    def exportar(self, usuario):
        self.client.force_login(usuario)
        r = self.client.get(reverse("expedientes:nomina_export"))
        self.assertEqual(r.status_code, 200)
        hoja = load_workbook(BytesIO(r.content)).active
        filas = list(hoja.values)
        encabezados = list(filas[0])
        datos = {f[0]: dict(zip(encabezados, f)) for f in filas[1:]}
        return hoja, encabezados, datos

    def columna(self, hoja, titulo):
        """Número de columna (1-based) de un encabezado."""
        for c in range(1, hoja.max_column + 1):
            if hoja.cell(row=1, column=c).value == titulo:
                return c
        raise AssertionError(f"no está la columna {titulo!r}")

    # --- Contenido -----------------------------------------------------------
    def test_estan_las_cuatro_columnas_bancarias(self):
        _, encabezados, _ = self.exportar(self.admin)
        for titulo in ("Banco", "Prefijo", "Número de cuenta", "Cuenta bancaria"):
            self.assertIn(titulo, encabezados)

    def test_trae_los_datos_de_cada_persona(self):
        _, _, datos = self.exportar(self.admin)
        fila = datos["V-1"]
        self.assertEqual(fila["Banco"], "Banco de Venezuela")
        self.assertEqual(fila["Prefijo"], "0102")
        self.assertEqual(fila["Número de cuenta"], "1234567890123456")

    def test_la_cuenta_completa_se_arma_sola(self):
        _, _, datos = self.exportar(self.admin)
        self.assertEqual(datos["V-1"]["Cuenta bancaria"], "01021234567890123456")

    def test_sin_datos_de_contratacion_las_celdas_van_vacias(self):
        """No puede salir 'None' ni un cero que parezca una cuenta."""
        _, _, datos = self.exportar(self.admin)
        fila = datos["V-2"]
        for titulo in ("Banco", "Prefijo", "Número de cuenta", "Cuenta bancaria"):
            self.assertIn(fila[titulo], ("", None), f"{titulo} trajo {fila[titulo]!r}")

    # --- El cero inicial -----------------------------------------------------
    def test_el_cero_inicial_del_prefijo_no_se_pierde(self):
        hoja, _, datos = self.exportar(self.admin)
        self.assertEqual(datos["V-1"]["Prefijo"], "0102")
        self.assertNotEqual(datos["V-1"]["Prefijo"], 102)
        self.assertIsInstance(datos["V-1"]["Prefijo"], str)

    def test_las_columnas_de_cuenta_quedan_con_formato_de_texto(self):
        """Aunque el valor esté bien, Excel reformatea si la columna es numérica."""
        hoja, _, _ = self.exportar(self.admin)
        for titulo in ("Prefijo", "Número de cuenta", "Cuenta bancaria", "C.I."):
            col = self.columna(hoja, titulo)
            for fila in range(2, hoja.max_row + 1):
                self.assertEqual(
                    hoja.cell(row=fila, column=col).number_format, "@",
                    f"{titulo} (fila {fila}) tendría que ser texto",
                )

    def test_la_cuenta_larga_no_sale_en_notacion_cientifica(self):
        """20 dígitos como número se muestran como 1,02346E+19."""
        _, _, datos = self.exportar(self.admin)
        valor = datos["V-1"]["Cuenta bancaria"]
        self.assertNotIn("E+", str(valor))
        self.assertEqual(len(str(valor)), 20)

    # --- Permisos ------------------------------------------------------------
    def test_solo_lectura_no_recibe_los_datos_bancarios(self):
        _, encabezados, datos = self.exportar(self.lectura_norte)
        for titulo in ("Banco", "Prefijo", "Número de cuenta", "Cuenta bancaria"):
            self.assertNotIn(titulo, encabezados)
        # Pero el listado le sigue saliendo, con las tallas.
        self.assertIn("V-1", datos)
        self.assertIn("Talla camisa", encabezados)

    def test_rrhh_interior_los_recibe_solo_de_su_zona(self):
        _, encabezados, datos = self.exportar(self.rrhh_norte)
        self.assertIn("Cuenta bancaria", encabezados)
        self.assertIn("V-1", datos)
        self.assertNotIn("V-2", datos)

    def test_queda_asentado_en_la_auditoria(self):
        from expedientes.models import RegistroAuditoria

        self.exportar(self.admin)
        ultima = RegistroAuditoria.objects.filter(entidad="Nómina").latest("id")
        self.assertIn("datos bancarios", ultima.descripcion)

    def test_para_solo_lectura_la_auditoria_no_dice_que_los_vio(self):
        from expedientes.models import RegistroAuditoria

        self.exportar(self.lectura_norte)
        ultima = RegistroAuditoria.objects.filter(entidad="Nómina").latest("id")
        self.assertNotIn("datos bancarios", ultima.descripcion)

    # --- Orden de las columnas ----------------------------------------------
    def test_van_agrupadas_y_antes_de_las_tallas(self):
        """Quien arma la transferencia quiere identidad y banco juntos."""
        _, encabezados, _ = self.exportar(self.admin)
        pos = {t: i for i, t in enumerate(encabezados)}
        self.assertLess(pos["Fecha de ingreso"], pos["Banco"])
        self.assertLess(pos["Banco"], pos["Prefijo"])
        self.assertLess(pos["Prefijo"], pos["Número de cuenta"])
        self.assertLess(pos["Número de cuenta"], pos["Cuenta bancaria"])
        self.assertLess(pos["Cuenta bancaria"], pos["Talla camisa"])
