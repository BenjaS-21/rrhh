"""Tests de la sección de hijos del expediente."""

from datetime import date, timedelta
from io import BytesIO

from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from expedientes.models import Hijo, Trabajador
from expedientes.tests import CLAVE, BasePagos


class EdadDelHijo(TestCase):
    """La edad se calcula, no se guarda: una cifra escrita queda vieja."""

    def hijo(self, nacimiento):
        return Hijo(nombre_completo="X", fecha_nacimiento=nacimiento)

    def test_cuenta_anios_cumplidos(self):
        hoy = timezone.localdate()
        self.assertEqual(self.hijo(hoy.replace(year=hoy.year - 10)).edad, 10)

    def test_el_dia_antes_del_cumple_todavia_no_los_cumplio(self):
        hoy = timezone.localdate()
        # Nació hace 10 años y un día menos: aún tiene 9.
        casi = hoy.replace(year=hoy.year - 10) + timedelta(days=1)
        self.assertEqual(self.hijo(casi).edad, 9)

    def test_el_dia_del_cumple_ya_los_cumplio(self):
        hoy = timezone.localdate()
        self.assertEqual(self.hijo(hoy.replace(year=hoy.year - 10)).edad, 10)

    def test_recien_nacido_tiene_cero(self):
        self.assertEqual(self.hijo(timezone.localdate()).edad, 0)

    def test_marca_a_los_menores(self):
        hoy = timezone.localdate()
        self.assertTrue(self.hijo(hoy.replace(year=hoy.year - 17)).es_menor)
        self.assertFalse(self.hijo(hoy.replace(year=hoy.year - 18)).es_menor)


class SeccionDeHijos(BasePagos):

    def url_agregar(self, trabajador=None):
        return reverse("expedientes:hijo_agregar",
                       args=[(trabajador or self.trab_norte).pk])

    def datos(self, **cambios):
        d = {"hijo-nombre_completo": "Maria Jose Perez",
             "hijo-fecha_nacimiento": "2015-03-12",
             "hijo-documento_identidad": "", "hijo-observaciones": ""}
        d.update(cambios)
        return d

    def agregar(self, usuario=None, trabajador=None, **cambios):
        self.client.force_login(usuario or self.admin)
        return self.client.post(self.url_agregar(trabajador), self.datos(**cambios))

    # --- Alta ----------------------------------------------------------------
    def test_se_agrega_y_queda_en_el_expediente(self):
        r = self.agregar()
        self.assertEqual(r.status_code, 302)
        hijo = Hijo.objects.get(trabajador=self.trab_norte)
        self.assertEqual(hijo.nombre_completo, "MARIA JOSE PEREZ")
        self.assertEqual(hijo.fecha_nacimiento, date(2015, 3, 12))
        self.assertEqual(hijo.creado_por, self.admin)

    def test_el_nombre_se_normaliza(self):
        """'ana  perez' y 'Ana Perez' tienen que ser el mismo."""
        self.agregar(**{"hijo-nombre_completo": "  ana   perez "})
        self.assertEqual(Hijo.objects.get().nombre_completo, "ANA PEREZ")

    def test_la_cantidad_sale_del_listado(self):
        self.agregar()
        self.agregar(**{"hijo-nombre_completo": "Carlos Perez",
                        "hijo-fecha_nacimiento": "2019-09-04"})
        self.trab_norte.refresh_from_db()
        self.assertEqual(self.trab_norte.cantidad_hijos, 2)

    def test_no_se_puede_cargar_dos_veces_al_mismo(self):
        """Doble clic en Agregar no puede duplicar al mismo chico."""
        self.agregar()
        self.client.force_login(self.admin)
        r = self.client.post(self.url_agregar(), self.datos(), follow=True)
        self.assertEqual(Hijo.objects.count(), 1)
        mensajes = " ".join(str(m) for m in r.context["messages"])
        self.assertIn("ya está cargado", mensajes)

    def test_rechaza_una_fecha_de_nacimiento_futura(self):
        manana = (timezone.localdate() + timedelta(days=1)).isoformat()
        r = self.agregar(**{"hijo-fecha_nacimiento": manana})
        self.assertEqual(r.status_code, 302)
        self.assertFalse(Hijo.objects.exists())

    def test_la_fecha_es_obligatoria(self):
        r = self.agregar(**{"hijo-fecha_nacimiento": ""})
        self.assertEqual(r.status_code, 302)
        self.assertFalse(Hijo.objects.exists())

    # --- Baja ----------------------------------------------------------------
    def test_se_puede_quitar(self):
        self.agregar()
        hijo = Hijo.objects.get()
        self.client.force_login(self.admin)
        r = self.client.post(reverse("expedientes:hijo_borrar", args=[hijo.pk]))
        self.assertEqual(r.status_code, 302)
        self.assertFalse(Hijo.objects.exists())

    def test_no_se_puede_quitar_por_GET(self):
        self.agregar()
        hijo = Hijo.objects.get()
        self.client.force_login(self.admin)
        r = self.client.get(reverse("expedientes:hijo_borrar", args=[hijo.pk]))
        self.assertEqual(r.status_code, 405)
        self.assertTrue(Hijo.objects.exists())

    def test_dar_de_baja_al_trabajador_se_lleva_a_los_hijos(self):
        self.agregar()
        self.trab_norte.delete()
        self.assertFalse(Hijo.objects.exists())

    # --- Permisos ------------------------------------------------------------
    def test_solo_lectura_los_ve_pero_no_los_carga(self):
        self.agregar()
        self.client.force_login(self.lectura_norte)
        cuerpo = self.client.get(self.url_detalle(self.trab_norte)).content.decode()
        self.assertIn("MARIA JOSE PEREZ", cuerpo)
        self.assertNotIn("Agregar hijo", cuerpo)

        r = self.client.post(self.url_agregar(), self.datos(
            **{"hijo-nombre_completo": "Otro"}))
        self.assertEqual(r.status_code, 403)
        self.assertEqual(Hijo.objects.count(), 1)

    def test_rrhh_de_otra_zona_no_puede_cargar(self):
        r = self.agregar(usuario=self.rrhh_sur)
        self.assertEqual(r.status_code, 403)
        self.assertFalse(Hijo.objects.exists())

    def test_rrhh_de_su_zona_si_puede(self):
        r = self.agregar(usuario=self.rrhh_norte)
        self.assertEqual(r.status_code, 302)
        self.assertTrue(Hijo.objects.exists())

    def test_no_se_puede_quitar_un_hijo_de_otra_zona(self):
        self.agregar()
        hijo = Hijo.objects.get()
        self.client.force_login(self.rrhh_sur)
        r = self.client.post(reverse("expedientes:hijo_borrar", args=[hijo.pk]))
        self.assertEqual(r.status_code, 403)
        self.assertTrue(Hijo.objects.exists())

    # --- Pantalla ------------------------------------------------------------
    def test_la_pantalla_muestra_la_cantidad_y_la_edad(self):
        self.agregar(**{"hijo-fecha_nacimiento":
                        timezone.localdate().replace(
                            year=timezone.localdate().year - 7).isoformat()})
        self.client.force_login(self.admin)
        cuerpo = self.client.get(self.url_detalle(self.trab_norte)).content.decode()
        self.assertIn("Hijos", cuerpo)
        self.assertIn("7 años", cuerpo)
        self.assertIn("menor", cuerpo)

    def test_sin_hijos_lo_dice(self):
        self.client.force_login(self.admin)
        cuerpo = self.client.get(self.url_detalle(self.trab_norte)).content.decode()
        self.assertIn("No hay hijos cargados", cuerpo)

    # --- Auditoría -----------------------------------------------------------
    def test_queda_registrado_el_alta_y_la_baja(self):
        from expedientes.models import RegistroAuditoria

        self.agregar()
        hijo = Hijo.objects.get()
        self.client.post(reverse("expedientes:hijo_borrar", args=[hijo.pk]))
        descripciones = list(
            RegistroAuditoria.objects.filter(entidad="Hijo")
            .values_list("descripcion", flat=True))
        self.assertTrue(any("Agregó a MARIA JOSE PEREZ" in d for d in descripciones))
        self.assertTrue(any("Quitó a MARIA JOSE PEREZ" in d for d in descripciones))


class HijosEnElExcel(BasePagos):

    def test_trae_la_cantidad(self):
        Hijo.objects.create(trabajador=self.trab_norte, nombre_completo="A",
                            fecha_nacimiento=date(2015, 1, 1))
        Hijo.objects.create(trabajador=self.trab_norte, nombre_completo="B",
                            fecha_nacimiento=date(2019, 1, 1))
        self.client.force_login(self.admin)
        r = self.client.get(reverse("expedientes:nomina_export"))
        hoja = load_workbook(BytesIO(r.content)).active
        filas = list(hoja.values)
        encabezados = list(filas[0])
        self.assertIn("Hijos", encabezados)
        datos = {f[0]: dict(zip(encabezados, f)) for f in filas[1:]}
        self.assertEqual(datos["V-1"]["Hijos"], 2)
        self.assertEqual(datos["V-2"]["Hijos"], 0)

    def test_solo_lectura_tambien_la_ve(self):
        """No es dato de pago: sigue la regla del resto del expediente."""
        self.client.force_login(self.lectura_norte)
        r = self.client.get(reverse("expedientes:nomina_export"))
        hoja = load_workbook(BytesIO(r.content)).active
        self.assertIn("Hijos", [c.value for c in hoja[1]])
        self.assertNotIn("Cuenta bancaria", [c.value for c in hoja[1]])

    def test_contar_los_hijos_no_agrega_consultas_por_trabajador(self):
        """El costo tiene que ser el mismo con 5 personas que con 20."""
        def poblar(desde, hasta):
            for i in range(desde, hasta):
                t = Trabajador.objects.create(
                    documento_identidad=f"V-90{i}", nombres="N", apellidos="A",
                    sede=self.sede_norte)
                Hijo.objects.create(trabajador=t, nombre_completo=f"H{i}",
                                    fecha_nacimiento=date(2015, 1, 1))

        self.client.force_login(self.admin)
        url = reverse("expedientes:nomina_export")

        poblar(0, 5)
        with CaptureQueriesContext(connection) as pocos:
            self.client.get(url)

        poblar(5, 20)
        with CaptureQueriesContext(connection) as muchos:
            self.client.get(url)

        self.assertEqual(len(muchos), len(pocos),
                         "el número de consultas crece con la cantidad de gente")
