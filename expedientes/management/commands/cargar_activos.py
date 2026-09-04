"""Carga la nómina activa desde el Excel del sistema viejo (ACTIVOS AL …).

    python manage.py cargar_activos "ACTIVOS AL 28082026 DIANA.xlsx"

Es idempotente: las cédulas que ya están se saltan. Todo se hace en una sola
transacción: o entra el archivo entero o no entra nada.

Decisiones de limpieza (validadas con Gestión Humana antes de la primera carga):
- Años de nacimiento imposibles (2081 en vez de 1981): se les resta 100.
- Emails repetidos en el archivo (rrhh@… y tienda.*@…, que son placeholders):
  no se cargan. Solo quedan los correos únicos de cada persona.
- Teléfonos en formato internacional (58 + número): se les quita el 58. Los
  rellenos de ceros y lo que no queda como un número venezolano válido, vacío.
- Estado civil "SIN DEFINIR" queda vacío; el resto se mapea a las opciones.
- El RIF sale de RIFTRAB con formato V-14097602-0.
- Nacionalidad distinta de Venezolana -> tipo de cédula E.
- Remuneración: la misma terna que recibieron todas las cargas anteriores
  (Sueldo base 130 Bs + Complemento de alimentación 260 $ + Ticket 40 $).
  El archivo no trae montos; se ajusta persona por persona en la ficha.
"""

import datetime
import re
import unicodedata

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from cuentas.models import (
    Cargo, Departamento, Sede, TipoDocumentoIdentidad, Zona,
)
from expedientes.models import (
    AsignacionPago, ConceptoPago, DatosContratacion, Trabajador,
)


def norm(t):
    t = unicodedata.normalize("NFD", str(t or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip().upper()


def cedula_de(v):
    if isinstance(v, float):
        v = int(v)
    return re.sub(r"\D", "", str(v or ""))


def fecha_de(v, hoy):
    """Date del Excel.

    Los años imposibles por error de tipeo (2081 en vez de 1981) bajan 100.
    Lo que siga siendo imposible para un trabajador (nacido después de 2010,
    o sea basura como 2026-01-01) queda vacío: fecha mala es peor que ninguna.
    """
    if not isinstance(v, (datetime.datetime, datetime.date)):
        return None
    d = v.date() if isinstance(v, datetime.datetime) else v
    if d.year > hoy.year:
        try:
            d = d.replace(year=d.year - 100)
        except ValueError:      # 29 de febrero en año no bisiesto
            d = d.replace(year=d.year - 100, day=28)
    if d.year > 2010:
        return None
    return d


def telefono_de(*valores):
    """El primer número que quede venezolano válido (11 dígitos, 04…)."""
    for v in valores:
        d = re.sub(r"\D", "", str(v or ""))
        if d.startswith("58") and len(d) >= 12:
            d = "0" + d[2:]
        if len(d) == 11 and d.startswith("04") and d[-7:] != "0000000":
            return d
    return ""


def rif_de(riftrab, cedula):
    """RIFTRAB -> 'V-14097602-0'.

    El sistema viejo no es consistente: unas filas traen V+cédula+dígito
    verificador y otras solo V+cédula. Se decide comparando con la cédula de
    la misma fila: si el resto calza, hay verificador; si no, el RIF va sin él.
    Lo que no calza de ninguna de las dos formas es basura y queda vacío.
    """
    t = re.sub(r"[^0-9A-Za-z]", "", str(riftrab or "")).upper()
    if len(t) < 3 or not t[0].isalpha() or not t[1:].isdigit() or not cedula:
        return ""
    sin_letra = t[1:]
    if sin_letra[:-1] == cedula or sin_letra[:-1] == cedula.zfill(8):
        return f"{t[0]}-{sin_letra[:-1]}-{sin_letra[-1]}"
    if sin_letra == cedula or sin_letra == cedula.zfill(8):
        return f"{t[0]}-{sin_letra}"
    return ""


EDO_CIVIL = {
    "SOLTERO": "SOLTERO(A)", "CASADO": "CASADO(A)",
    "DIVORCIADO": "DIVORCIADO(A)", "VIUDO": "VIUDO(A)",
    "CONCUBINO": "CONCUBINO(A)",
}

# Lo que dice el archivo -> sede. Los que no son renombres directos pasan por
# la regla general (se les quita "TIENDA DAMASCO" y los romanos se arábigos).
ALIAS_TIENDAS = {
    "SEDE CORPORATIVA": "CORPORACION",
    "CENDIS GUATIRE I": "CENTRO DE DISTRIBUCION GUATIRE I",
    "CENDIS GUATIRE II": "CENTRO DE DISTRIBUCION GUATIRE II",
    "CENDIS GUATIRE III": "CENTRO DE DISTRIBUCION GUATIRE III",
    "CENDIS GUATIRE IV": "CENTRO DE DISTRIBUCION GUATIRE IV",
    "CENDIS LA YAGUARA": "YAGUARA",
    "TIENDA DAMASCO VALLE LA PASCUA": "VALLE DE PASCUA",
    "TIENDA DAMASCO LA CALIFORNIA": "CALIFORNIA",
    "TIENDA DAMASCO LA TRINIDAD": "TRINIDAD",
    "TIENDA DAMASCO LA CANDELARIA": "CANDELARIA",
    "TIENDA DAMASCO MERIDA": "MERIDAD",
    "TIENDA DAMASCO CATIA ANTUAN": "ANTUAN",
    "TIENDA DAMASCO CATIA MUEBLERIA": "MUEBLERIA",
    "TIENDA DAMASCO C.C.C.T.": "CCCT",
}

# Sedes que el sistema no tenía y hacen falta para este archivo.
SEDES_NUEVAS = {
    "EXCLUSIVE": "MIRANDA",
    "PALO NEGRO": "ARAGUA",
    "PUERTO AYACUCHO": "AMAZONAS",
    "CENTRO DE DISTRIBUCION GUATIRE IV": "MIRANDA",
}

ROMAN = {"I": "1", "II": "2", "III": "3", "IV": "4", "V": "5"}


class Command(BaseCommand):
    help = "Carga la nómina activa desde el Excel del sistema viejo."

    def add_arguments(self, parser):
        parser.add_argument("archivo", help="Ruta del Excel de activos.")

    # -- resolución de catálogos ------------------------------------------
    def _sedes(self):
        sedes = {norm(s.nombre): s for s in Sede.objects.select_related("zona")}
        for nombre, zona_nombre in SEDES_NUEVAS.items():
            if norm(nombre) in sedes:
                continue
            zona, _ = Zona.objects.get_or_create(nombre=zona_nombre)
            sede = Sede.objects.create(nombre=nombre, zona=zona)
            sedes[norm(nombre)] = sede
            self.stdout.write(f"  sede creada: {nombre} ({zona_nombre})")
        return sedes

    @staticmethod
    def _candidatos_tienda(tienda):
        t = ALIAS_TIENDAS.get(tienda, tienda)
        t = re.sub(r"^(TIENDA\s+DAMASCO|DAMASCO|SEDE|CENDIS)\s+", "", t).strip()
        candidatos = [t]
        m = re.search(r" (I|II|III|IV|V)$", t)
        if m:
            candidatos.append(t[:m.start()] + " " + ROMAN[m.group(1)])
            candidatos.append(t[:m.start()].strip())
        return candidatos

    # -- comando -----------------------------------------------------------
    def handle(self, *args, **options):
        ruta = options["archivo"]
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        hoja = wb[wb.sheetnames[0]]
        filas = list(hoja.iter_rows(min_row=2, values_only=True))
        hoy = datetime.date.today()

        repetidos_email = set()
        vistos_email = set()
        for f in filas:
            e = str(f[19] or "").strip().lower()
            if e:
                (repetidos_email if e in vistos_email else vistos_email).add(e)

        tipo_v = TipoDocumentoIdentidad.objects.get(codigo="V")
        tipo_e = TipoDocumentoIdentidad.objects.get(codigo="E")
        # La terna de remuneración de todas las cargas anteriores. El archivo
        # no trae montos; se ajusta persona por persona en la ficha.
        conceptos = [(ConceptoPago.objects.get(pk=1), "130.00"),
                     (ConceptoPago.objects.get(pk=3), "260.00"),
                     (ConceptoPago.objects.get(pk=4), "40.00")]
        sedes = self._sedes()
        unidades = {norm(d.nombre): d for d in Departamento.objects.all()}
        catalogo = {
            (norm(c.nombre), norm(c.departamento.nombre)): c
            for c in Cargo.objects.select_related("departamento")
        }
        # Cualquier fila con ese nombre sirve: se prefiere la general.
        por_nombre = {}
        for c in sorted(catalogo.values(),
                        key=lambda c: (not c.es_general, c.pk)):
            por_nombre.setdefault(norm(c.nombre), c)
        existentes = set(Trabajador.objects.values_list("documento_identidad", flat=True))

        cargados = saltados = 0
        anios_corregidos = emails_omitidos = telefonos_vacios = 0
        cargos_creados = set()

        with transaction.atomic():
            for f in filas:
                cedula = cedula_de(f[3])
                if not cedula:
                    continue
                if cedula in existentes:
                    saltados += 1
                    continue

                tienda = norm(f[33])
                sede = next((sedes[c] for c in self._candidatos_tienda(tienda)
                             if c in sedes), None)
                if sede is None:
                    raise CommandError(f"Sede sin resolver: {tienda!r} (ced {cedula})")

                unidad = unidades.get(norm(f[28]))
                if unidad is None:
                    raise CommandError(f"Unidad sin resolver: {norm(f[28])!r} (ced {cedula})")

                nombre_cargo = norm(f[29])
                cargo = catalogo.get((nombre_cargo, norm(f[28])))
                if cargo is None:
                    # Si el nombre ya existe en cualquier unidad, se usa ese:
                    # crear otra fila con el mismo nombre es justo lo que el
                    # catálogo general vino a terminar.
                    cargo = por_nombre.get(nombre_cargo)
                if cargo is None:
                    cargo = Cargo.objects.create(nombre=nombre_cargo,
                                                 departamento=unidad)
                    por_nombre[nombre_cargo] = cargo
                    cargos_creados.add(nombre_cargo)
                catalogo[(nombre_cargo, norm(f[28]))] = cargo

                nacimiento = fecha_de(f[22], hoy)
                if isinstance(f[22], (datetime.datetime, datetime.date)) \
                        and nacimiento and f[22].year != nacimiento.year:
                    anios_corregidos += 1

                email = str(f[19] or "").strip().lower()
                if email and email in repetidos_email:
                    emails_omitidos += 1
                    email = ""

                telefono = telefono_de(f[17], f[18])
                if not telefono:
                    telefonos_vacios += 1

                es_venezolano = norm(f[8]) in ("", "VENEZOLANA", "VENEZOLANO")
                nombres = " ".join(x for x in (norm(f[4]), norm(f[5])) if x)
                apellidos = " ".join(x for x in (norm(f[6]), norm(f[7])) if x)

                t = Trabajador.objects.create(
                    tipo_documento=tipo_v if es_venezolano else tipo_e,
                    documento_identidad=cedula,
                    rif=rif_de(f[9], cedula),
                    nombres=nombres, apellidos=apellidos,
                    fecha_nacimiento=nacimiento,
                    email=email, telefono=telefono,
                    sede=sede, departamento=unidad, puesto=cargo,
                    fecha_ingreso=fecha_de(f[27], hoy),
                )
                DatosContratacion.objects.create(
                    trabajador=t,
                    estado_civil=EDO_CIVIL.get(norm(f[21]), ""),
                    direccion=str(f[11] or "").strip()[:400],
                    ciudad_nacimiento=norm(f[23])[:150],
                )
                for concepto, monto in conceptos:
                    AsignacionPago.objects.create(
                        trabajador=t, concepto=concepto, monto=monto,
                        moneda=concepto.moneda)
                existentes.add(cedula)
                cargados += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nCargados: {cargados} | saltados (ya estaban): {saltados}"
            f"\n  años de nacimiento corregidos: {anios_corregidos}"
            f"\n  emails omitidos (genéricos/repetidos): {emails_omitidos}"
            f"\n  sin teléfono válido: {telefonos_vacios}"
        ))
        if cargos_creados:
            self.stdout.write(f"  cargos creados: {sorted(cargos_creados)}")
