"""Tests de la remuneración del expediente.

Cubren tres cosas: que los conceptos del catálogo bajen como ítems a completar,
que los montos se registren en la moneda correcta (sin conversión), y que todo
respete el rol y la zona del usuario.
"""

import re
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.db import IntegrityError, transaction
from django.db.models import ProtectedError
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from cuentas.models import Cargo, Departamento, Sede, Zona
from expedientes.models import (
    AsignacionPago, ConceptoPago, DatosContratacion, Moneda, Trabajador,
)

Usuario = get_user_model()

CLAVE = "Clave-De-Prueba-123"


class BasePagos(TestCase):
    """Base de casi todo. Con la restricción por zona PRENDIDA.

    De fábrica el sistema no restringe por zona: todos ven todo. Estos tests
    son justamente los que ejercitan el modo restringido, así que lo prenden a
    propósito. Los que verifican el comportamiento de fábrica lo apagan y lo
    dicen en su nombre.
    """

    @classmethod
    def setUpTestData(cls):
        cls.restringir_por_zona(True)
        cls.norte = Zona.objects.create(nombre="Norte")
        cls.sur = Zona.objects.create(nombre="Sur")
        cls.sede_norte = Sede.objects.create(nombre="Salta", zona=cls.norte)
        cls.sede_sur = Sede.objects.create(nombre="Neuquén", zona=cls.sur)

        cls.trab_norte = Trabajador.objects.create(
            documento_identidad="V-1", nombres="Ana", apellidos="Norte",
            sede=cls.sede_norte, fecha_ingreso=date(2020, 3, 15),
        )
        cls.trab_sur = Trabajador.objects.create(
            documento_identidad="V-2", nombres="Beto", apellidos="Sur",
            sede=cls.sede_sur, fecha_ingreso=date(2022, 11, 1),
        )

        # Vienen de la migración de datos 0004.
        cls.bs = Moneda.objects.get(codigo="VES")
        cls.usd = Moneda.objects.get(codigo="USD")
        cls.eur = Moneda.objects.get(codigo="EUR")
        cls.sueldo = ConceptoPago.objects.get(nombre="Sueldo base")

        # Catálogo mínimo de cargos: el campo dejó de ser texto libre.
        cls.unidad = Departamento.objects.create(nombre="TIENDA DE PRUEBA")
        cls.cargo_cajera = Cargo.objects.create(nombre="CAJERA", departamento=cls.unidad)
        cls.cargo_supervisora = Cargo.objects.create(
            nombre="SUPERVISORA", departamento=cls.unidad)

        cls.admin = cls._usuario("admin_nac", Usuario.Rol.ADMIN, None)
        cls.rrhh_norte = cls._usuario("rrhh_norte", Usuario.Rol.RRHH_INTERIOR, cls.norte)
        cls.rrhh_sur = cls._usuario("rrhh_sur", Usuario.Rol.RRHH_INTERIOR, cls.sur)
        cls.lectura_norte = cls._usuario("lectura_norte", Usuario.Rol.SOLO_LECTURA, cls.norte)

    @classmethod
    def _usuario(cls, username, rol, zona):
        u = Usuario.objects.create_user(username=username, password=CLAVE)
        u.rol = rol
        u.zona = zona
        u.save()
        return u

    # --- Atajos ------------------------------------------------------------
    def url_detalle(self, trabajador):
        return reverse("expedientes:trabajador_detail", args=[trabajador.pk])

    def url_grilla(self, trabajador):
        return reverse("expedientes:remuneracion_guardar", args=[trabajador.pk])

    def url_extra(self, trabajador):
        return reverse("expedientes:pago_agregar", args=[trabajador.pk])

    def campo(self, concepto):
        return f"rem-concepto_{concepto.pk}"

    def datos_extra(self, **cambios):
        datos = {
            "extra-nombre_libre": "Bono puntualidad",
            "extra-monto": "400.00",
            "extra-moneda": self.usd.pk,
            "extra-observaciones": "",
        }
        datos.update(cambios)
        return datos


    @classmethod
    def restringir_por_zona(cls, activo):
        from configuracion.models import Preferencias

        Preferencias.objects.update_or_create(
            pk=1, defaults={"restringir_por_zona": activo})


class GrillaDeRemuneracion(BasePagos):
    """Los conceptos del catálogo bajan como ítems con su casillero de monto."""

    def setUp(self):
        self.client.force_login(self.admin)
        self.bono_usd = ConceptoPago.objects.create(
            nombre="Bono producción", clase=ConceptoPago.Clase.BONO,
            moneda=self.usd, orden=20,
        )

    def test_lista_todos_los_conceptos_activos_como_items(self):
        resp = self.client.get(self.url_detalle(self.trab_norte))
        nombres = [c.nombre for c, _campo in resp.context["form_remuneracion"].filas()]
        self.assertEqual(nombres, ["Sueldo base", "Bono producción"])

    def test_un_concepto_nuevo_del_catalogo_aparece_solo(self):
        ConceptoPago.objects.create(nombre="Cesta ticket", moneda=self.bs, orden=30)
        resp = self.client.get(self.url_detalle(self.trab_norte))
        nombres = [c.nombre for c, _ in resp.context["form_remuneracion"].filas()]
        self.assertIn("Cesta ticket", nombres)

    def test_un_concepto_inactivo_no_aparece(self):
        self.bono_usd.activo = False
        self.bono_usd.save()
        resp = self.client.get(self.url_detalle(self.trab_norte))
        nombres = [c.nombre for c, _ in resp.context["form_remuneracion"].filas()]
        self.assertNotIn("Bono producción", nombres)

    def test_guarda_solo_los_montos_completados(self):
        self.client.post(self.url_grilla(self.trab_norte), {
            self.campo(self.sueldo): "180.00",
            self.campo(self.bono_usd): "",  # esta persona no cobra este concepto
        })
        pagos = AsignacionPago.objects.filter(trabajador=self.trab_norte, activo=True)
        self.assertEqual(pagos.count(), 1)
        self.assertEqual(pagos.first().concepto, self.sueldo)

    def test_la_moneda_la_define_el_concepto(self):
        self.client.post(self.url_grilla(self.trab_norte), {
            self.campo(self.sueldo): "180.00",
            self.campo(self.bono_usd): "400.00",
        })
        por_concepto = {
            p.concepto_id: p
            for p in AsignacionPago.objects.filter(trabajador=self.trab_norte, activo=True)
        }
        self.assertEqual(por_concepto[self.sueldo.pk].moneda, self.bs)
        self.assertEqual(por_concepto[self.bono_usd.pk].moneda, self.usd)

    def test_precarga_los_montos_vigentes(self):
        AsignacionPago.objects.create(trabajador=self.trab_norte, concepto=self.sueldo,
                                      monto=Decimal("180"), moneda=self.bs)
        resp = self.client.get(self.url_detalle(self.trab_norte))
        form = resp.context["form_remuneracion"]
        self.assertEqual(form.initial[f"concepto_{self.sueldo.pk}"], Decimal("180"))

    def test_cambiar_el_monto_actualiza_la_misma_fila(self):
        pago = AsignacionPago.objects.create(trabajador=self.trab_norte, concepto=self.sueldo,
                                             monto=Decimal("180"), moneda=self.bs)
        self.client.post(self.url_grilla(self.trab_norte),
                         {self.campo(self.sueldo): "220.50"})
        pago.refresh_from_db()
        self.assertEqual(pago.monto, Decimal("220.50"))
        self.assertEqual(AsignacionPago.objects.filter(trabajador=self.trab_norte).count(), 1)

    def test_vaciar_el_monto_da_de_baja_el_concepto(self):
        pago = AsignacionPago.objects.create(trabajador=self.trab_norte, concepto=self.sueldo,
                                             monto=Decimal("180"), moneda=self.bs)
        self.client.post(self.url_grilla(self.trab_norte), {self.campo(self.sueldo): ""})
        pago.refresh_from_db()
        self.assertFalse(pago.activo)

    def test_rechaza_monto_negativo(self):
        self.client.post(self.url_grilla(self.trab_norte), {self.campo(self.sueldo): "-5"})
        self.assertEqual(AsignacionPago.objects.count(), 0)

    def test_no_toca_los_bonos_extras(self):
        extra = AsignacionPago.objects.create(
            trabajador=self.trab_norte, nombre_libre="Bono puntualidad",
            monto=Decimal("50"), moneda=self.bs,
        )
        self.client.post(self.url_grilla(self.trab_norte), {self.campo(self.sueldo): "180"})
        extra.refresh_from_db()
        self.assertTrue(extra.activo)

    def test_sin_conceptos_en_el_catalogo_la_grilla_queda_vacia(self):
        AsignacionPago.objects.all().delete()
        ConceptoPago.objects.all().delete()
        resp = self.client.get(self.url_detalle(self.trab_norte))
        self.assertEqual(resp.context["form_remuneracion"].filas(), [])
        self.assertContains(resp, "Creálos en Configuración")


class BonosExtras(BasePagos):
    """Bonos puntuales que no están en el catálogo."""

    def setUp(self):
        self.client.force_login(self.admin)

    def test_agrega_bono_con_nombre_libre(self):
        resp = self.client.post(self.url_extra(self.trab_norte), self.datos_extra())
        self.assertEqual(resp.status_code, 302)

        pago = AsignacionPago.objects.get(trabajador=self.trab_norte)
        self.assertIsNone(pago.concepto)
        self.assertTrue(pago.es_bono_extra)
        self.assertEqual(pago.etiqueta, "Bono puntualidad")
        self.assertEqual(pago.monto_formateado, "400,00 $")
        self.assertEqual(pago.creado_por, self.admin)

    def test_el_monto_en_divisa_no_se_convierte(self):
        """400 $ se guarda como 400 en USD, no como su equivalente en Bs."""
        self.client.post(self.url_extra(self.trab_norte), self.datos_extra())
        pago = AsignacionPago.objects.get(trabajador=self.trab_norte)
        self.assertEqual(pago.monto, Decimal("400.00"))
        self.assertEqual(pago.moneda.codigo, "USD")

    def test_exige_nombre(self):
        self.client.post(self.url_extra(self.trab_norte),
                         self.datos_extra(**{"extra-nombre_libre": "  "}))
        self.assertEqual(AsignacionPago.objects.count(), 0)

    def test_rechaza_monto_cero_o_negativo(self):
        for monto in ("0", "-5"):
            self.client.post(self.url_extra(self.trab_norte),
                             self.datos_extra(**{"extra-monto": monto}))
        self.assertEqual(AsignacionPago.objects.count(), 0)

    def test_quitar_es_baja_logica(self):
        self.client.post(self.url_extra(self.trab_norte), self.datos_extra())
        pago = AsignacionPago.objects.get(trabajador=self.trab_norte)

        self.client.post(reverse("expedientes:pago_borrar", args=[pago.pk]))
        pago.refresh_from_db()
        self.assertFalse(pago.activo)
        self.assertTrue(AsignacionPago.objects.filter(pk=pago.pk).exists())

    def test_los_del_catalogo_no_se_editan_por_esa_pantalla(self):
        pago = AsignacionPago.objects.create(trabajador=self.trab_norte, concepto=self.sueldo,
                                             monto=Decimal("180"), moneda=self.bs)
        resp = self.client.get(reverse("expedientes:pago_editar", args=[pago.pk]))
        self.assertRedirects(resp, self.url_detalle(self.trab_norte))


class TotalesYFormato(BasePagos):
    """Las monedas se totalizan por separado: no hay tasa de cambio."""

    def setUp(self):
        self.client.force_login(self.admin)

    def test_totales_se_agrupan_por_moneda(self):
        for nombre, monto, moneda in [
            ("Sueldo", "180.00", self.bs),
            ("Bono A", "400.00", self.usd),
            ("Bono B", "400.00", self.usd),
            ("Bono C", "50.00", self.eur),
        ]:
            AsignacionPago.objects.create(
                trabajador=self.trab_norte, nombre_libre=nombre,
                monto=Decimal(monto), moneda=moneda,
            )

        resp = self.client.get(self.url_detalle(self.trab_norte))
        totales = {t["moneda"].codigo: t["total"] for t in resp.context["totales_pagos"]}
        self.assertEqual(totales, {
            "VES": "180,00 Bs",
            "USD": "800,00 $",
            "EUR": "50,00 €",
        })

    def test_mismo_trabajador_admite_varias_monedas(self):
        """El caso del pedido: 180 Bs de sueldo y 400 $ de bono, a la vez."""
        AsignacionPago.objects.create(trabajador=self.trab_norte, concepto=self.sueldo,
                                      monto=Decimal("180"), moneda=self.bs)
        AsignacionPago.objects.create(trabajador=self.trab_norte, nombre_libre="Bono",
                                      monto=Decimal("400"), moneda=self.usd)
        self.assertEqual(self.trab_norte.pagos.filter(activo=True).count(), 2)

    def test_formato_de_montos_grandes(self):
        self.assertEqual(self.bs.formatear(Decimal("1234567.5")), "1.234.567,50 Bs")

    def test_la_base_impide_concepto_y_nombre_libre_juntos(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            AsignacionPago.objects.create(
                trabajador=self.trab_norte, concepto=self.sueldo,
                nombre_libre="Bono suelto", monto=Decimal("1"), moneda=self.bs,
            )


class MonedaDelConcepto(BasePagos):
    """Cada concepto del catálogo define en qué moneda se paga."""

    def setUp(self):
        self.client.force_login(self.admin)
        self.bono_usd = ConceptoPago.objects.create(
            nombre="Bono producción", clase=ConceptoPago.Clase.BONO, moneda=self.usd,
        )

    def test_el_concepto_tiene_moneda(self):
        # La migración le asignó la nacional al concepto que ya existía.
        self.assertEqual(self.sueldo.moneda, self.bs)
        self.assertEqual(self.bono_usd.moneda, self.usd)

    def test_el_catalogo_muestra_la_moneda(self):
        resp = self.client.get(reverse("configuracion:lista", args=["conceptos-pago"]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Moneda")
        self.assertContains(resp, "Bs (VES)")
        self.assertContains(resp, "$ (USD)")

    def test_no_se_puede_borrar_una_moneda_en_uso(self):
        with self.assertRaises(ProtectedError):
            self.usd.delete()


class PermisosDeRemuneracion(BasePagos):
    """Los montos los ven Admin y RRHH Interior de su zona. Solo lectura no."""

    def test_admin_ve_la_seccion(self):
        self.client.force_login(self.admin)
        resp = self.client.get(self.url_detalle(self.trab_sur))
        self.assertTrue(resp.context["puede_pagos"])
        self.assertContains(resp, "Remuneración")

    def test_rrhh_ve_la_seccion_en_su_zona(self):
        self.client.force_login(self.rrhh_norte)
        resp = self.client.get(self.url_detalle(self.trab_norte))
        self.assertTrue(resp.context["puede_pagos"])

    def test_solo_lectura_no_ve_la_seccion(self):
        self.client.force_login(self.lectura_norte)
        resp = self.client.get(self.url_detalle(self.trab_norte))
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(resp.context["puede_pagos"])
        self.assertNotContains(resp, "Remuneración")

    def test_solo_lectura_no_puede_guardar_la_grilla(self):
        self.client.force_login(self.lectura_norte)
        resp = self.client.post(self.url_grilla(self.trab_norte),
                                {self.campo(self.sueldo): "180"})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(AsignacionPago.objects.count(), 0)

    def test_solo_lectura_no_puede_agregar_bonos(self):
        self.client.force_login(self.lectura_norte)
        resp = self.client.post(self.url_extra(self.trab_norte), self.datos_extra())
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(AsignacionPago.objects.count(), 0)

    def test_rrhh_de_otra_zona_no_puede_guardar_la_grilla(self):
        self.client.force_login(self.rrhh_sur)
        resp = self.client.post(self.url_grilla(self.trab_norte),
                                {self.campo(self.sueldo): "180"})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(AsignacionPago.objects.count(), 0)

    def test_rrhh_de_otra_zona_no_puede_agregar_bonos(self):
        self.client.force_login(self.rrhh_sur)
        resp = self.client.post(self.url_extra(self.trab_norte), self.datos_extra())
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(AsignacionPago.objects.count(), 0)

    def test_rrhh_de_otra_zona_no_puede_editar_ni_borrar(self):
        pago = AsignacionPago.objects.create(
            trabajador=self.trab_norte, nombre_libre="Bono",
            monto=Decimal("10"), moneda=self.bs,
        )
        self.client.force_login(self.rrhh_sur)
        self.assertEqual(
            self.client.get(reverse("expedientes:pago_editar", args=[pago.pk])).status_code, 403
        )
        self.assertEqual(
            self.client.post(reverse("expedientes:pago_borrar", args=[pago.pk])).status_code, 403
        )
        pago.refresh_from_db()
        self.assertTrue(pago.activo)

    def test_rrhh_sin_zona_no_accede_con_la_restriccion_prendida(self):
        """Restringido por zona y sin zona: no ve ningún monto."""
        huerfano = self._usuario("sin_zona", Usuario.Rol.RRHH_INTERIOR, None)
        self.client.force_login(huerfano)
        resp = self.client.post(self.url_extra(self.trab_norte), self.datos_extra())
        self.assertEqual(resp.status_code, 403)

    def test_sin_restriccion_un_rrhh_sin_zona_si_accede(self):
        """Cómo funciona de fábrica: sin restricción, cualquiera carga a cualquiera.

        Vale la pena escribirlo: es la contracara de todos los tests de "otra
        zona no puede" que hay más arriba, y deja claro que esos dependen de una
        opción, no de una barrera que esté siempre puesta.
        """
        self.restringir_por_zona(False)
        huerfano = self._usuario("sin_zona2", Usuario.Rol.RRHH_INTERIOR, None)
        self.client.force_login(huerfano)
        resp = self.client.post(self.url_extra(self.trab_norte), self.datos_extra())
        self.assertEqual(resp.status_code, 302)

    def test_anonimo_va_al_login(self):
        resp = self.client.post(self.url_extra(self.trab_norte), self.datos_extra())
        self.assertEqual(resp.status_code, 302)
        self.assertIn("ingresar", resp["Location"])


class AltaUnificada(BasePagos):
    """El alta pide en una sola pantalla todo lo que necesitan los documentos."""

    def setUp(self):
        self.client.force_login(self.admin)

    def datos_alta(self, **cambios):
        datos = {
            # Ficha
            "documento_identidad": "V-99", "nombres": "ANA MARIA",
            "apellidos": "LOPEZ PEREZ", "fecha_nacimiento": "1995-06-10",
            "email": "", "telefono": "0412-1234567",
            "sede": self.sede_norte.pk, "departamento": "",
            "puesto": self.cargo_cajera.pk, "fecha_ingreso": "2026-09-01",
            # Contratación
            "contrato-estado_civil": "SOLTERO(A)",
            "contrato-direccion": "CALLE 5, CASA 12, MARACAY",
            "contrato-ciudad_nacimiento": "MARACAY, ARAGUA",
            "contrato-duracion_dias": "90",
            "contrato-fecha_culminacion": "",
            "contrato-motivo_contratacion": "Temporada Navidad",
            "contrato-horario": "8:00AM a 5:00PM",
            "contrato-ciudad_firma": "CARACAS",
            "contrato-banco": "Banco de Venezuela",
            "contrato-prefijo": "0102",
            "contrato-numero_cuenta": "1234567890123456",
            "contrato-observaciones": "", "contrato-responsable": "",
        }
        datos.update(cambios)
        return datos

    def test_crea_ficha_y_datos_de_contratacion_juntos(self):
        resp = self.client.post(reverse("expedientes:trabajador_create"),
                                self.datos_alta())
        self.assertEqual(resp.status_code, 302)

        t = Trabajador.objects.get(documento_identidad="V-99")
        self.assertEqual(t.puesto, self.cargo_cajera)
        self.assertEqual(t.contratacion.estado_civil, "SOLTERO(A)")
        self.assertEqual(t.contratacion.ciudad_nacimiento, "MARACAY, ARAGUA")
        self.assertEqual(t.contratacion.banco, "BANCO DE VENEZUELA")

    def test_la_duracion_calcula_la_fecha_de_fin(self):
        self.client.post(reverse("expedientes:trabajador_create"), self.datos_alta())
        datos = Trabajador.objects.get(documento_identidad="V-99").contratacion
        self.assertEqual(datos.fecha_culminacion, date(2026, 9, 1) + timedelta(days=90))

    def test_la_fecha_de_fin_calcula_la_duracion(self):
        self.client.post(reverse("expedientes:trabajador_create"), self.datos_alta(**{
            "contrato-duracion_dias": "",
            "contrato-fecha_culminacion": "2026-12-01",
        }))
        datos = Trabajador.objects.get(documento_identidad="V-99").contratacion
        self.assertEqual(datos.fecha_culminacion, date(2026, 12, 1))
        self.assertEqual(datos.duracion_dias, 91)

    def test_si_vienen_las_dos_manda_la_fecha(self):
        """La fecha es la que se imprime en el contrato."""
        self.client.post(reverse("expedientes:trabajador_create"), self.datos_alta(**{
            "contrato-duracion_dias": "5",              # inconsistente a propósito
            "contrato-fecha_culminacion": "2026-12-01",
        }))
        datos = Trabajador.objects.get(documento_identidad="V-99").contratacion
        self.assertEqual(datos.fecha_culminacion, date(2026, 12, 1))
        self.assertEqual(datos.duracion_dias, 91)

    def test_rechaza_fin_anterior_al_ingreso(self):
        resp = self.client.post(reverse("expedientes:trabajador_create"), self.datos_alta(**{
            "contrato-duracion_dias": "",
            "contrato-fecha_culminacion": "2026-01-01",
        }))
        self.assertEqual(resp.status_code, 200)  # vuelve al formulario
        self.assertFalse(Trabajador.objects.filter(documento_identidad="V-99").exists())

    def test_la_cuenta_bancaria_se_arma_sola(self):
        self.client.post(reverse("expedientes:trabajador_create"), self.datos_alta())
        datos = Trabajador.objects.get(documento_identidad="V-99").contratacion
        self.assertEqual(datos.cuenta_bancaria, "01021234567890123456")

    def test_rechaza_cuenta_con_letras(self):
        self.client.post(reverse("expedientes:trabajador_create"),
                         self.datos_alta(**{"contrato-numero_cuenta": "12AB-34"}))
        self.assertFalse(Trabajador.objects.filter(documento_identidad="V-99").exists())

    def test_el_responsable_por_defecto_es_quien_carga(self):
        self.client.post(reverse("expedientes:trabajador_create"), self.datos_alta())
        datos = Trabajador.objects.get(documento_identidad="V-99").contratacion
        self.assertEqual(datos.responsable, self.admin.get_username())

    def test_el_alta_no_pregunta_el_estado(self):
        resp = self.client.get(reverse("expedientes:trabajador_create"))
        self.assertNotIn("estado", resp.context["form"].fields)
        # Y al editar sí aparece, para poder dar de baja.
        self.client.post(reverse("expedientes:trabajador_create"), self.datos_alta())
        t = Trabajador.objects.get(documento_identidad="V-99")
        self.assertEqual(t.estado, Trabajador.Estado.ACTIVO)
        resp = self.client.get(reverse("expedientes:trabajador_update", args=[t.pk]))
        self.assertIn("estado", resp.context["form"].fields)

    def test_todos_los_campos_quedan_en_alguna_seccion(self):
        """Si se agrega un campo y no se ubica, cae en 'Otros datos', no desaparece.

        Se abren las parejas —el tipo de cédula y la cédula comparten casilla—
        porque compartir casilla no los hace un campo solo.
        """
        from expedientes.views import _uno_por_uno

        resp = self.client.get(reverse("expedientes:trabajador_create"))
        en_pantalla = {
            campo.name
            for seccion in resp.context["secciones"]
            for campo in _uno_por_uno(seccion["campos"])
        }
        esperados = (set(resp.context["form"].fields)
                     | set(resp.context["form_contrato"].fields))
        self.assertEqual(en_pantalla, esperados)

    def test_las_secciones_estan_en_orden(self):
        resp = self.client.get(reverse("expedientes:trabajador_create"))
        titulos = [s["titulo"] for s in resp.context["secciones"]]
        self.assertEqual(titulos, ["Datos personales", "Puesto y contrato",
                                   "Datos bancarios", "Tallas de uniforme",
                                   "Seguimiento"])

    def test_no_pide_los_campos_derivados(self):
        """Edad, día/mes/año y dirección de tienda se calculan, no se piden."""
        resp = self.client.get(reverse("expedientes:trabajador_create"))
        campos = set(resp.context["form"].fields) | set(resp.context["form_contrato"].fields)
        for derivado in ["edad", "dia_nacimiento", "mes_nacimiento", "ano_nacimiento",
                         "direccion_tienda", "cuenta_bancaria", "salario"]:
            self.assertNotIn(derivado, campos)

    def test_editar_conserva_los_datos_de_contratacion(self):
        self.client.post(reverse("expedientes:trabajador_create"), self.datos_alta())
        t = Trabajador.objects.get(documento_identidad="V-99")

        self.client.post(reverse("expedientes:trabajador_update", args=[t.pk]),
                         self.datos_alta(**{"puesto": self.cargo_supervisora.pk,
                                            "estado": Trabajador.Estado.ACTIVO}))
        t.refresh_from_db()
        self.assertEqual(t.puesto, self.cargo_supervisora)
        self.assertEqual(t.contratacion.ciudad_nacimiento, "MARACAY, ARAGUA")
        self.assertEqual(DatosContratacion.objects.filter(trabajador=t).count(), 1)

    def test_rrhh_de_otra_zona_no_puede_editar(self):
        self.client.post(reverse("expedientes:trabajador_create"), self.datos_alta())
        t = Trabajador.objects.get(documento_identidad="V-99")
        self.client.force_login(self.rrhh_sur)
        resp = self.client.post(reverse("expedientes:trabajador_update", args=[t.pk]),
                                self.datos_alta())
        self.assertEqual(resp.status_code, 403)


class ExportNomina(BasePagos):
    """El Excel lleva fecha de ingreso y montos, respetando quién puede verlos."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        AsignacionPago.objects.create(trabajador=cls.trab_norte, concepto=cls.sueldo,
                                      monto=Decimal("180.00"), moneda=cls.bs)
        AsignacionPago.objects.create(trabajador=cls.trab_norte, nombre_libre="Bono producción",
                                      monto=Decimal("400.00"), moneda=cls.usd)
        AsignacionPago.objects.create(trabajador=cls.trab_sur, concepto=cls.sueldo,
                                      monto=Decimal("250.00"), moneda=cls.bs)

    def _exportar(self):
        resp = self.client.get(reverse("expedientes:nomina_export"))
        self.assertEqual(resp.status_code, 200)
        hoja = load_workbook(BytesIO(resp.content)).active
        filas = list(hoja.values)
        encabezados = list(filas[0])
        # Una fila por trabajador, indexada por C.I. para no depender del orden.
        datos = {f[0]: dict(zip(encabezados, f)) for f in filas[1:]}
        return encabezados, datos

    def test_incluye_fecha_de_ingreso(self):
        self.client.force_login(self.admin)
        encabezados, datos = self._exportar()
        self.assertIn("Fecha de ingreso", encabezados)

        valor = datos["V-1"]["Fecha de ingreso"]
        self.assertEqual(valor.date() if hasattr(valor, "date") else valor,
                         date(2020, 3, 15))

    def test_una_columna_por_concepto_con_su_moneda(self):
        """El encabezado usa el nombre del catálogo, no la moneda sola."""
        self.client.force_login(self.admin)
        encabezados, datos = self._exportar()
        self.assertIn("Sueldo base (Bs)", encabezados)
        self.assertEqual(float(datos["V-1"]["Sueldo base (Bs)"]), 180.0)
        self.assertEqual(float(datos["V-2"]["Sueldo base (Bs)"]), 250.0)

    def test_no_hay_totales_por_moneda(self):
        self.client.force_login(self.admin)
        encabezados, _ = self._exportar()
        self.assertFalse([h for h in encabezados if str(h).startswith("Total")])
        self.assertNotIn("Detalle de pagos", encabezados)

    def test_un_concepto_nuevo_del_catalogo_agrega_su_columna(self):
        ConceptoPago.objects.create(nombre="Bono de transporte", moneda=self.usd)
        self.client.force_login(self.admin)
        encabezados, datos = self._exportar()
        self.assertIn("Bono de transporte ($)", encabezados)
        # Nadie lo cobra todavía: la celda va vacía, no en cero.
        self.assertIsNone(datos["V-1"]["Bono de transporte ($)"])

    def test_concepto_que_no_cobra_deja_la_celda_vacia(self):
        otro = ConceptoPago.objects.create(nombre="Bono nocturno", moneda=self.bs)
        AsignacionPago.objects.create(trabajador=self.trab_norte, concepto=otro,
                                      monto=Decimal("75.00"), moneda=self.bs)
        self.client.force_login(self.admin)
        _, datos = self._exportar()
        self.assertEqual(float(datos["V-1"]["Bono nocturno (Bs)"]), 75.0)
        self.assertIsNone(datos["V-2"]["Bono nocturno (Bs)"])

    def test_un_concepto_desactivado_con_pagos_vigentes_conserva_su_columna(self):
        """Desactivar el concepto no puede hacer desaparecer plata ya asignada."""
        self.sueldo.activo = False
        self.sueldo.save(update_fields=["activo"])
        self.addCleanup(lambda: ConceptoPago.objects.filter(pk=self.sueldo.pk)
                        .update(activo=True))
        self.client.force_login(self.admin)
        encabezados, datos = self._exportar()
        self.assertIn("Sueldo base (Bs)", encabezados)
        self.assertEqual(float(datos["V-1"]["Sueldo base (Bs)"]), 180.0)

    def test_solo_lectura_no_recibe_columnas_de_montos(self):
        self.client.force_login(self.lectura_norte)
        encabezados, datos = self._exportar()
        self.assertIn("Fecha de ingreso", encabezados)
        self.assertNotIn("Sueldo base (Bs)", encabezados)
        self.assertNotIn("Detalle de pagos", encabezados)
        # El listado sigue saliendo, solo que sin la parte salarial.
        self.assertIn("V-1", datos)

    def test_rrhh_interior_exporta_montos_solo_de_su_zona(self):
        self.client.force_login(self.rrhh_norte)
        encabezados, datos = self._exportar()
        self.assertIn("Sueldo base (Bs)", encabezados)
        self.assertIn("V-1", datos)
        self.assertNotIn("V-2", datos)  # trabajador de la zona Sur


class TallasUniformeTests(BasePagos):
    """Tallas de dotación: se cargan en la ficha, se ven y se exportan."""

    def test_seccion_de_tallas_aparece_en_el_formulario(self):
        self.client.login(username="admin_nac", password=CLAVE)
        r = self.client.get(
            reverse("expedientes:trabajador_update", args=[self.trab_norte.pk])
        )
        self.assertContains(r, "Tallas de uniforme")
        for campo in ("talla_camisa", "talla_pantalon", "talla_zapato"):
            self.assertContains(r, f"id_contrato-{campo}")

    def test_se_guardan_normalizadas_en_mayuscula(self):
        self.client.login(username="admin_nac", password=CLAVE)
        r = self.client.post(
            reverse("expedientes:trabajador_update", args=[self.trab_norte.pk]),
            {
                "documento_identidad": self.trab_norte.documento_identidad,
                "nombres": "Ana", "apellidos": "Norte",
                "sede": self.sede_norte.pk, "estado": "ACTIVO",
                "contrato-talla_camisa": " m ",
                "contrato-talla_pantalon": "32",
                "contrato-talla_zapato": "41",
            },
        )
        self.assertEqual(r.status_code, 302)
        datos = DatosContratacion.objects.get(trabajador=self.trab_norte)
        self.assertEqual(datos.talla_camisa, "M")
        self.assertEqual(datos.talla_pantalon, "32")
        self.assertEqual(datos.talla_zapato, "41")
        self.assertEqual(datos.tallas, "Camisa M · Pantalón 32 · Zapato 41")

    def test_tallas_vacias_dan_resumen_vacio(self):
        datos = DatosContratacion.objects.create(trabajador=self.trab_sur)
        self.assertEqual(datos.tallas, "")

    def test_detalle_muestra_las_tallas(self):
        DatosContratacion.objects.create(
            trabajador=self.trab_norte, talla_camisa="L", talla_zapato="43",
        )
        self.client.login(username="admin_nac", password=CLAVE)
        r = self.client.get(self.url_detalle(self.trab_norte))
        self.assertContains(r, "Camisa L")
        self.assertContains(r, "Zapato 43")

    def test_excel_trae_las_columnas_de_talla(self):
        DatosContratacion.objects.create(
            trabajador=self.trab_norte, talla_camisa="M",
            talla_pantalon="32", talla_zapato="41",
        )
        self.client.login(username="admin_nac", password=CLAVE)
        r = self.client.get(reverse("expedientes:nomina_export"))
        self.assertEqual(r.status_code, 200)

        hoja = load_workbook(BytesIO(r.content)).active
        encabezados = [c.value for c in hoja[1]]
        for titulo in ("Talla camisa", "Talla pantalón", "Talla zapato"):
            self.assertIn(titulo, encabezados)

        col_ci = encabezados.index("C.I.") + 1
        fila = next(f for f in range(2, hoja.max_row + 1)
                    if hoja.cell(row=f, column=col_ci).value == "V-1")
        self.assertEqual(hoja.cell(row=fila,
                                   column=encabezados.index("Talla camisa") + 1).value, "M")
        self.assertEqual(hoja.cell(row=fila,
                                   column=encabezados.index("Talla zapato") + 1).value, "41")

    def test_solo_lectura_tambien_ve_las_columnas_de_talla(self):
        """Las tallas no son dato salarial: no dependen de `puede_editar`."""
        DatosContratacion.objects.create(
            trabajador=self.trab_norte, talla_camisa="M",
        )
        self.client.login(username="lectura_norte", password=CLAVE)
        r = self.client.get(reverse("expedientes:nomina_export"))
        hoja = load_workbook(BytesIO(r.content)).active
        encabezados = [c.value for c in hoja[1]]
        self.assertIn("Talla camisa", encabezados)
        self.assertNotIn("Detalle de pagos", encabezados)


class FechasEnElFormulario(BasePagos):
    """Regresión: al editar, las fechas ya cargadas no se pueden borrar solas.

    `<input type="date">` solo entiende `AAAA-MM-DD`. Si el valor se renderiza
    con el formato local (`05/03/1990`), el navegador muestra el campo vacío y
    al guardar la fecha se pierde sin que nadie lo pida.
    """

    def setUp(self):
        self.trab_norte.fecha_nacimiento = date(1990, 3, 5)
        self.trab_norte.save(update_fields=["fecha_nacimiento"])
        DatosContratacion.objects.update_or_create(
            trabajador=self.trab_norte,
            defaults={"fecha_culminacion": date(2026, 10, 31)},
        )
        self.client.login(username="admin_nac", password=CLAVE)

    def test_las_fechas_se_renderizan_en_iso(self):
        r = self.client.get(
            reverse("expedientes:trabajador_update", args=[self.trab_norte.pk])
        )
        cuerpo = r.content.decode()
        for esperado in ('value="1990-03-05"', 'value="2020-03-15"',
                         'value="2026-10-31"'):
            self.assertIn(esperado, cuerpo)
        # El formato local en un input date deja el campo vacío en el navegador.
        for prohibido in ("05/03/1990", "15/03/2020", "31/10/2026"):
            self.assertNotIn(f'value="{prohibido}"', cuerpo)

    def test_abrir_editar_y_guardar_no_borra_las_fechas(self):
        """Simula el navegador: un input date descarta el valor que no sabe leer.

        Es lo que pasaba de verdad: Django escribía `value="05/03/1990"`, el
        campo se veía vacío y al guardar llegaba "" -> la fecha se borraba.
        """
        url = reverse("expedientes:trabajador_update", args=[self.trab_norte.pk])
        cuerpo = self.client.get(url).content.decode()

        iso = re.compile(r"^\d{4}-\d{2}-\d{2}$")
        datos = {}
        for etiqueta in re.findall(r"<(?:input|select|textarea)\s[^>]*>", cuerpo):
            nombre = re.search(r'name="([^"]+)"', etiqueta)
            if not nombre or nombre.group(1) == "csrfmiddlewaretoken":
                continue
            valor = re.search(r'value="([^"]*)"', etiqueta)
            valor = valor.group(1) if valor else ""
            # El navegador vacía el campo si el valor no es AAAA-MM-DD.
            if 'type="date"' in etiqueta and valor and not iso.match(valor):
                valor = ""
            datos[nombre.group(1)] = valor

        # Los <select> no traen value en la etiqueta de apertura.
        datos["sede"] = str(self.sede_norte.pk)
        datos["estado"] = self.trab_norte.estado

        r = self.client.post(url, datos)
        self.assertEqual(r.status_code, 302, _errores(r))

        self.trab_norte.refresh_from_db()
        self.assertEqual(self.trab_norte.fecha_nacimiento, date(1990, 3, 5))
        self.assertEqual(self.trab_norte.fecha_ingreso, date(2020, 3, 15))
        self.assertEqual(
            DatosContratacion.objects.get(trabajador=self.trab_norte).fecha_culminacion,
            date(2026, 10, 31),
        )


def _errores(respuesta):
    """Errores de los formularios, para que un 200 inesperado diga por qué."""
    partes = []
    for clave in ("form", "form_contrato"):
        form = (respuesta.context or {}).get(clave)
        if form is not None and form.errors:
            partes.append(f"{clave}: {form.errors.as_json()}")
    return " | ".join(partes) or "sin errores de formulario"


class FiltroPorVariasTiendas(BasePagos):
    """El filtro de tiendas acepta más de una a la vez."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        # Tres tiendas en la zona Norte, una persona en cada una.
        cls.sede_b = Sede.objects.create(nombre="Jujuy", zona=cls.norte)
        cls.sede_c = Sede.objects.create(nombre="Tucumán", zona=cls.norte)
        cls.trab_b = Trabajador.objects.create(
            documento_identidad="V-10", nombres="Caro", apellidos="Bravo",
            sede=cls.sede_b, fecha_ingreso=date(2021, 1, 4),
        )
        cls.trab_c = Trabajador.objects.create(
            documento_identidad="V-11", nombres="Dante", apellidos="Cruz",
            sede=cls.sede_c, fecha_ingreso=date(2021, 6, 9),
        )

    def cedulas(self, respuesta):
        return {t.documento_identidad for t in respuesta.context["pagina"]}

    def listar(self, usuario=None, **filtros):
        self.client.force_login(usuario or self.admin)
        return self.client.get(reverse("expedientes:trabajador_list"), filtros)

    def test_dos_tiendas_traen_las_dos(self):
        r = self.listar(sedes=[self.sede_norte.pk, self.sede_b.pk])
        self.assertEqual(self.cedulas(r), {"V-1", "V-10"})

    def test_tres_tiendas_traen_las_tres(self):
        r = self.listar(sedes=[self.sede_norte.pk, self.sede_b.pk, self.sede_c.pk])
        self.assertEqual(self.cedulas(r), {"V-1", "V-10", "V-11"})

    def test_una_sola_tienda_sigue_funcionando(self):
        r = self.listar(sedes=[self.sede_b.pk])
        self.assertEqual(self.cedulas(r), {"V-10"})

    def test_sin_tiendas_marcadas_salen_todos(self):
        r = self.listar()
        self.assertEqual(self.cedulas(r), {"V-1", "V-2", "V-10", "V-11"})

    def test_un_enlace_viejo_con_sede_sigue_andando(self):
        """Los favoritos y el historial usan `?sede=`, de cuando era una sola."""
        r = self.listar(sede=self.sede_b.pk)
        self.assertEqual(self.cedulas(r), {"V-10"})

    def test_se_combina_con_la_busqueda_por_texto(self):
        r = self.listar(sedes=[self.sede_norte.pk, self.sede_b.pk], q="Caro")
        self.assertEqual(self.cedulas(r), {"V-10"})

    def test_rrhh_interior_no_ve_tiendas_de_otra_zona(self):
        """Aunque fuerce el id de una tienda ajena, no puede espiar esa zona."""
        r = self.listar(self.rrhh_norte, sedes=[self.sede_sur.pk])
        self.assertNotIn("V-2", self.cedulas(r))

    def test_el_desplegable_solo_ofrece_tiendas_de_su_zona(self):
        r = self.listar(self.rrhh_norte)
        ofrecidas = {s.nombre for s in r.context["form"].fields["sedes"].queryset}
        self.assertEqual(ofrecidas, {"Salta", "Jujuy", "Tucumán"})

    def test_el_resumen_dice_cuantas_hay_marcadas(self):
        form = self.listar(sedes=[self.sede_norte.pk, self.sede_b.pk]).context["form"]
        self.assertEqual(form.resumen_tiendas, "2 tiendas")

        form = self.listar(sedes=[self.sede_b.pk]).context["form"]
        self.assertEqual(form.resumen_tiendas, "Jujuy")

        self.assertEqual(self.listar().context["form"].resumen_tiendas,
                         "Todas las tiendas")

    def test_la_nomina_filtra_por_varias_tiendas(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("expedientes:nomina"),
                            {"sedes": [self.sede_b.pk, self.sede_c.pk]})
        self.assertEqual(self.cedulas(r), {"V-10", "V-11"})

    def test_el_excel_respeta_el_filtro_de_varias_tiendas(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("expedientes:nomina_export"),
                            {"sedes": [self.sede_b.pk, self.sede_c.pk]})
        hoja = load_workbook(BytesIO(r.content)).active
        cedulas = {f[0] for f in list(hoja.values)[1:]}
        self.assertEqual(cedulas, {"V-10", "V-11"})

    def test_la_paginacion_conserva_los_filtros(self):
        """Pasar de página no puede devolverte al listado completo."""
        r = self.listar(sedes=[self.sede_norte.pk, self.sede_b.pk], q="a")
        querystring = r.context["querystring"]
        self.assertIn(f"sedes={self.sede_norte.pk}", querystring)
        self.assertIn(f"sedes={self.sede_b.pk}", querystring)
        self.assertIn("q=a", querystring)
        self.assertNotIn("page=", querystring)
